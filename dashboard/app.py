"""
Dashboard completo com dados reais do mercado - Todas as funcionalidades consolidadas
Sistema Multi-Tenant com autenticação
"""
from flask import Flask, render_template, jsonify, request as flask_request, send_file, make_response, redirect, url_for
from flask_login import LoginManager, login_required, current_user
import io
import sys
import os
from datetime import datetime
import pandas as pd
import logging
import traceback
import secrets

# Adicionar diretórios ao path
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, base_dir)

# Carregar variáveis de ambiente do arquivo .env (se existir)
try:
    from dotenv import load_dotenv
    env_path = os.path.join(base_dir, '.env')
    load_dotenv(env_path)
except ImportError:
    # python-dotenv não instalado, continuar sem ele
    pass

# Configurar logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Importar módulos reais
try:
    from data.price_collector import coletar_precos, coletar_ultimo_pregao
    from data.trades_repository import (
        init_db as trades_init_db, import_csv as trades_import_csv, list_trades as trades_list, 
        positions_summary, reset_trades, calculate_realized_pnl,
        import_dividendos_csv_bytes, import_dividendos_excel_bytes, list_dividendos, 
        calculate_total_dividendos, reset_dividendos
    )
    from data.dividendos_collector import sincronizar_dividendos_automatico, coletar_dividendos_brapi
    from data.trades_repository import verificar_necessidade_sincronizacao_dividendos
    from data.news_collector import coletar_noticias_brasileiras
    from features.technical_indicators import calcular_todos_indicadores
    from features.statistical_features import calcular_todas_features_estatisticas
    from features.sentiment_engine import analisar_sentimento_noticias
    from strategies.trend_strategy import gerar_sinal_tendencia
    from strategies.mean_reversion_strategy import gerar_sinal_reversao
    from strategies.news_strategy import gerar_sinal_noticias
    from core.signal_orchestrator import SignalOrchestrator
    from core.trade_executor import TradeExecutor
    from core.risk_manager import RiskManager
    from core.ia_advisor import analisar_carteira_com_ia
    from core.investment_advisor import analisar_carteira_completa
    from utils.config import TICKERS, PESOS_ESTRATEGIAS, CAPITAL_INICIAL
    MODULOS_CARREGADOS = True
except ImportError as e:
    print(f"⚠️  Aviso: Alguns módulos não foram carregados: {e}")
    logger.warning(f"Módulos não carregados: {e}")
    MODULOS_CARREGADOS = False
    TICKERS = ['BBSE3.SA', 'CMIG4.SA', 'CSMG3.SA', 'ITUB4.SA', 'PETR4.SA', 'SANB11.SA', 'SYN3.SA']
    PESOS_ESTRATEGIAS = {'trend': 0.4, 'reversao': 0.3, 'news': 0.3}
    CAPITAL_INICIAL = 10000.0

# Configurar Flask para encontrar templates
template_dir = os.path.join(os.path.dirname(__file__), 'templates')
app = Flask(__name__, template_folder=template_dir)

# Configurar Flask-Login
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'auth.login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'

# Importar modelo de usuário
from auth.models import User, init_auth_db

@login_manager.user_loader
def load_user(user_id):
    """Callback para carregar usuário da sessão"""
    return User.get(int(user_id))

# Registrar blueprint de autenticação
from auth.auth_routes import auth_bp
app.register_blueprint(auth_bp)

# Estado do sistema
try:
    executor = TradeExecutor(modo_mock=True) if MODULOS_CARREGADOS else None
    # Inicializar base de dados de trades e autenticação
    try:
        trades_init_db()
        init_auth_db()
        logger.info("🗄️  Base de dados de operações e autenticação inicializada.")
    except Exception as e:
        logger.warning(f"⚠️  Não foi possível inicializar DB: {e}")
except:
    executor = None

capital_atual = CAPITAL_INICIAL

@app.route('/')
@login_required
def index():
    """Home - Importar operações e visualizar carteira"""
    try:
        # Tentar usar home.html primeiro (mais completo), fallback para index.html
        try:
            return render_template('home.html')
        except:
            return render_template('index.html')
    except Exception as e:
        error_msg = f"Erro ao carregar template: {str(e)}\n{traceback.format_exc()}"
        logger.error(f"❌ Erro na rota principal: {error_msg}")
        return f"""
        <html>
        <head><title>Erro - Dashboard</title></head>
        <body>
            <h1>Erro ao carregar dashboard</h1>
            <p>{error_msg}</p>
            <p>Verifique se o arquivo home.html ou index.html existe em dashboard/templates/</p>
        </body>
        </html>
        """, 500

@app.route('/api/status')
@login_required
def get_status():
    """Status do sistema"""
    try:
        capital_inicial = CAPITAL_INICIAL if 'CAPITAL_INICIAL' in globals() else 10000.0
        retorno = ((capital_atual - capital_inicial) / capital_inicial * 100) if capital_inicial > 0 else 0
        
        posicoes_abertas = 0
        total_operacoes = 0
        
        if executor is not None:
            try:
                posicoes = executor.obter_posicoes_abertas()
                posicoes_abertas = len(posicoes) if posicoes else 0
                
                historico = executor.obter_historico()
                total_operacoes = len(historico) if not historico.empty else 0
            except:
                pass
        
        return jsonify({
            'capital_atual': capital_atual,
            'capital_inicial': capital_inicial,
            'retorno_percentual': retorno,
            'posicoes_abertas': posicoes_abertas,
            'total_operacoes': total_operacoes,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        logger.error(f"❌ Erro ao obter status: {e}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/analisar/<ticker>')
@login_required
def analisar_ticker(ticker):
    """Análise REAL com dados do mercado"""
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados. Verifique as dependências.'}), 500
        
        logger.info(f"\n{'='*70}")
        logger.info(f"🔍 [DASHBOARD] Iniciando análise para {ticker}")
        logger.info(f"{'='*70}")
        logger.info(f"📡 [DASHBOARD] Coletando dados reais de preços para {ticker}...")
        
        try:
            # 1) Buscar último pregão (modo leve para evitar 429)
            df_ultimo = coletar_ultimo_pregao(ticker)
            logger.info(f"✅ [DASHBOARD] Último pregão coletado: {len(df_ultimo)} período")
            
            # 2) Tentar expandir para um período maior; se falhar, seguir em modo mínimo
            modo_minimo = False
            try:
                df_precos = coletar_precos(ticker, periodo='1mo', intervalo='1d')
                logger.info(f"✅ [DASHBOARD] Período expandido coletado: {len(df_precos)} períodos")
            except Exception as e_expand:
                logger.warning(f"⚠️  [DASHBOARD] Não foi possível coletar período de 1 mês (seguindo com último pregão): {e_expand}")
                df_precos = df_ultimo.copy()
                modo_minimo = True
        except Exception as e:
            logger.error(f"❌ [DASHBOARD] Erro ao coletar preços: {e}")
            return jsonify({
                'erro': f'Erro ao coletar preços de {ticker}',
                'detalhe': str(e),
                'sugestao': 'Verifique se o ticker está correto e se há dados disponíveis no yfinance'
            }), 404
        
        if df_precos.empty:
            logger.error(f"❌ [DASHBOARD] DataFrame de preços está vazio para {ticker}")
            return jsonify({
                'erro': f'Nenhum dado encontrado para {ticker}',
                'sugestao': 'O ticker pode estar incorreto ou não ter dados disponíveis no período solicitado'
            }), 404
        
        # 2. Coletar notícias REAIS
        logger.info(f"📰 [DASHBOARD] Coletando notícias...")
        noticias = []
        try:
            noticias = coletar_noticias_brasileiras()
            logger.info(f"✅ [DASHBOARD] Notícias coletadas: {len(noticias)}")
        except Exception as e:
            logger.warning(f"⚠️  [DASHBOARD] Erro ao coletar notícias: {e}")
        
        # 3. Calcular indicadores REAIS
        df = df_precos.copy()
        if len(df_precos) > 1:
            logger.info(f"📊 [DASHBOARD] Calculando indicadores técnicos...")
            try:
                df = calcular_todos_indicadores(df_precos)
                logger.info(f"✅ [DASHBOARD] Indicadores técnicos calculados")
            except Exception as e:
                logger.error(f"❌ [DASHBOARD] Erro ao calcular indicadores técnicos: {e}")
                traceback.print_exc()
        
        logger.info(f"📈 [DASHBOARD] Calculando features estatísticas...")
        try:
            if len(df) > 1:
                df = calcular_todas_features_estatisticas(df)
                logger.info(f"✅ [DASHBOARD] Features estatísticas calculadas")
        except Exception as e:
            logger.warning(f"⚠️  [DASHBOARD] Erro ao calcular features estatísticas: {e}")
        
        # 4. Gerar sinais REAIS
        logger.info(f"🎯 [DASHBOARD] Gerando sinais das estratégias...")
        sinal_trend = pd.Series(0, index=df.index)
        sinal_reversao = pd.Series(0, index=df.index)
        sinal_news_valor = 0
        sentimento = 0.0
        
        try:
            if len(df) > 1:
                sinal_trend = gerar_sinal_tendencia(df)
        except Exception as e:
            logger.error(f"❌ Erro na estratégia de tendência: {e}")
        
        try:
            if len(df) > 1:
                sinal_reversao = gerar_sinal_reversao(df)
        except Exception as e:
            logger.error(f"❌ Erro na estratégia de reversão: {e}")
        
        try:
            sentimento = analisar_sentimento_noticias(noticias)
            sinal_news_valor = gerar_sinal_noticias(noticias)
        except Exception as e:
            logger.error(f"❌ Erro na estratégia de notícias: {e}")
        
        sinal_news = pd.Series([sinal_news_valor] * len(df), index=df.index)
        
        # 5. Combinar sinais REAIS
        logger.info(f"🎼 [DASHBOARD] Combinando sinais (orquestração)...")
        try:
            orchestrator = SignalOrchestrator(pesos=PESOS_ESTRATEGIAS)
            sinal_final = orchestrator.combinar_sinais(sinal_trend, sinal_reversao, sinal_news)
            confianca = orchestrator.calcular_confianca(sinal_trend, sinal_reversao, sinal_news)
        except Exception as e:
            logger.error(f"❌ Erro na orquestração: {e}")
            traceback.print_exc()
            sinal_final = pd.Series(0, index=df.index)
            confianca = pd.Series(0.5, index=df.index)
        
        # 6. Dados REAIS do último período
        ultimo_preco = float(df['Close'].iloc[-1])
        rsi = float(df['RSI'].iloc[-1]) if 'RSI' in df.columns and pd.notna(df['RSI'].iloc[-1]) else None
        macd = float(df['MACD'].iloc[-1]) if 'MACD' in df.columns and pd.notna(df['MACD'].iloc[-1]) else None
        ultimo_sinal = int(sinal_final.iloc[-1]) if len(sinal_final) > 0 else 0
        ultima_confianca = float(confianca.iloc[-1]) if len(confianca) > 0 else 0.5
        
        # 7. Variação do preço (último vs anterior)
        if len(df) >= 2:
            preco_anterior = float(df['Close'].iloc[-2])
            variacao_percentual = ((ultimo_preco - preco_anterior) / preco_anterior) * 100
        else:
            variacao_percentual = 0.0
        
        logger.info(f"✅ [DASHBOARD] Análise concluída para {ticker}")
        
        return jsonify({
            'ticker': ticker,
            'preco_atual': ultimo_preco,
            'variacao_percentual': variacao_percentual,
            'sinal_final': ultimo_sinal,
            'confianca': ultima_confianca,
            'sentimento_noticias': float(sentimento),
            'rsi': rsi,
            'macd': macd,
            'sinal_trend': int(sinal_trend.iloc[-1]) if len(sinal_trend) > 0 else 0,
            'sinal_reversao': int(sinal_reversao.iloc[-1]) if len(sinal_reversao) > 0 else 0,
            'sinal_news': int(sinal_news_valor),
            'timestamp': datetime.now().isoformat(),
            'fonte_dados': 'MERCADO REAL',
            'modo_minimo': modo_minimo if 'modo_minimo' in locals() else False
        })
    
    except Exception as e:
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        logger.error(f"Erro completo: {error_msg}")
        return jsonify({'erro': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/operacoes')
@login_required
def get_operacoes():
    """Histórico de operações"""
    try:
        if executor is None:
            return jsonify({'operacoes': []})
        
        historico = executor.obter_historico()
        
        if historico.empty:
            return jsonify({'operacoes': []})
        
        operacoes = historico.to_dict('records')
        for op in operacoes:
            if 'timestamp' in op and pd.notna(op['timestamp']):
                op['timestamp'] = op['timestamp'].isoformat() if hasattr(op['timestamp'], 'isoformat') else str(op['timestamp'])
        
        return jsonify({'operacoes': operacoes[-10:]})
    except Exception as e:
        logger.error(f"❌ Erro ao obter operações: {e}")
        return jsonify({'operacoes': [], 'erro': str(e)})

@app.route('/api/posicoes')
@login_required
def get_posicoes():
    """Retorna posições abertas"""
    try:
        if executor is None:
            return jsonify({'posicoes': {}})
        
        posicoes = executor.obter_posicoes_abertas()
        
        return jsonify({'posicoes': posicoes})
    except Exception as e:
        logger.error(f"❌ Erro ao obter posições: {e}")
        return jsonify({'posicoes': {}, 'erro': str(e)})

@app.route('/api/importar_operacoes', methods=['POST'])
@login_required
def importar_operacoes():
    """Importa operações via Excel (preferencial) ou CSV (fallback).
    Campos esperados: date, ticker, side (BUY/SELL), quantity, price, fees
    """
    try:
        if 'file' not in flask_request.files:
            return jsonify({'erro': 'Arquivo não encontrado no formulário (campo "file")'}), 400
        file = flask_request.files['file']
        if file.filename == '':
            return jsonify({'erro': 'Nome de arquivo inválido'}), 400
        filename = file.filename.lower()
        data_bytes = file.read()
        # Preferir Excel
        if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            from data.trades_repository import import_excel_bytes, insert_rows
            import pandas as pd
            from io import BytesIO
            # Ler Excel e converter para rows
            df = pd.read_excel(BytesIO(data_bytes))
            rows = df.to_dict('records')
            resultado = insert_rows(rows, user_id=current_user.id)
        elif filename.endswith('.csv'):
            from data.trades_repository import import_csv_bytes, insert_rows
            import pandas as pd
            from io import BytesIO
            # Ler CSV e converter para rows
            df = pd.read_csv(BytesIO(data_bytes))
            rows = df.to_dict('records')
            resultado = insert_rows(rows, user_id=current_user.id)
        else:
            # Tentar detectar pelo conteúdo: se começar com PK é zip/xlsx
            from data.trades_repository import insert_rows
            import pandas as pd
            from io import BytesIO
            if data_bytes[:2] == b'PK':
                df = pd.read_excel(BytesIO(data_bytes))
            else:
                df = pd.read_csv(BytesIO(data_bytes))
            rows = df.to_dict('records')
            resultado = insert_rows(rows, user_id=current_user.id)
        return jsonify({'status': 'ok', 'inseridos': resultado.get('inserted', 0)})
    except Exception as e:
        logger.error(f"❌ Erro ao importar operações: {e}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/trades', methods=['GET'])
@login_required
def listar_trades():
    """Lista operações armazenadas (últimas 200)."""
    try:
        rows = trades_list(user_id=current_user.id)
        return jsonify({'trades': rows})
    except Exception as e:
        logger.error(f"❌ Erro ao listar trades: {e}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/portfolio_resumo', methods=['GET'])
@login_required
def portfolio_resumo():
    """Consolida posições, custos médios e avalia pelo último pregão."""
    try:
        resumo = positions_summary()
        pos = resumo.get('positions', [])
        total_investido = 0.0
        total_valor = 0.0
        detalhado = []
        hoje = datetime.now().date()
        
        for p in pos:
            ticker = p['ticker']
            qty = float(p['net_quantity'])
            avg_cost = float(p['avg_cost'])
            first_buy_date_str = p.get('first_buy_date')
            total_investido += max(qty, 0) * avg_cost
            
            # Tentar buscar preço, mas continuar mesmo se falhar
            preco = None
            preco_disponivel = False
            try:
                preco_df = coletar_ultimo_pregao(ticker)
                if not preco_df.empty:
                    preco = float(preco_df['Close'].iloc[-1])
                    preco_disponivel = True
            except Exception as e_preco:
                logger.warning(f"⚠️  Não foi possível obter preço para {ticker}: {str(e_preco)[:100]}")
                preco = None
                preco_disponivel = False
            
            # Calcular valor da posição (0 se preço não disponível)
            valor_posicao = qty * (preco if preco is not None else 0.0)
            if preco is not None:
                total_valor += valor_posicao
            
            # Calcular rentabilidade simples e anualizada
            rentabilidade_simples = None
            rentabilidade_anualizada = None
            if preco is not None and avg_cost > 0:
                try:
                    rentabilidade_simples = (preco - avg_cost) / avg_cost
                    
                    if first_buy_date_str:
                        try:
                            first_buy_date = datetime.strptime(first_buy_date_str, "%Y-%m-%d").date()
                            dias = (hoje - first_buy_date).days
                            
                            if dias > 0:
                                rentabilidade_anualizada = ((1 + rentabilidade_simples) ** (365.0 / dias)) - 1
                            elif dias == 0:
                                rentabilidade_anualizada = rentabilidade_simples
                        except Exception as e_rent:
                            logger.warning(f"⚠️  Erro ao calcular rentabilidade anualizada para {ticker}: {str(e_rent)[:100]}")
                except Exception as e_rent_simples:
                    logger.warning(f"⚠️  Erro ao calcular rentabilidade simples para {ticker}: {str(e_rent_simples)[:100]}")
            
            detalhado.append({
                'ticker': ticker,
                'quantidade': qty,
                'preco_medio': avg_cost,
                'preco_ultimo': preco if preco is not None else None,
                'preco_disponivel': preco_disponivel,
                'valor_posicao': valor_posicao if preco is not None else None,
                'rentabilidade': rentabilidade_simples,
                'rentabilidade_anualizada': rentabilidade_anualizada,
                'first_buy_date': first_buy_date_str
            })
        
        # Calcular PnL realizado (vendas executadas)
        pnl_realizado_info = calculate_realized_pnl(user_id=current_user.id)
        pnl_realizado = pnl_realizado_info.get('total_pnl_realizado', 0.0)
        custo_vendas = pnl_realizado_info.get('total_custo_vendas', 0.0)
        receita_vendas = pnl_realizado_info.get('total_receita_vendas', 0.0)
        
        # Sincronizar dividendos automaticamente (em background, não bloqueia resposta)
        try:
            tickers_para_sincronizar = [p['ticker'] for p in pos if verificar_necessidade_sincronizacao_dividendos(p['ticker'], user_id=current_user.id)]
            if tickers_para_sincronizar:
                logger.info(f"🔄 [PORTFOLIO] Sincronizando dividendos para {len(tickers_para_sincronizar)} tickers em background...")
                import threading
                def sync_bg():
                    try:
                        sincronizar_dividendos_automatico(tickers=tickers_para_sincronizar, forcar_atualizacao=False, user_id=current_user.id)
                    except Exception as e:
                        logger.warning(f"⚠️  [PORTFOLIO] Erro ao sincronizar dividendos em background: {str(e)[:100]}")
                threading.Thread(target=sync_bg, daemon=True).start()
        except Exception as e:
            logger.warning(f"⚠️  [PORTFOLIO] Erro ao verificar sincronização de dividendos: {str(e)[:100]}")
        
        # Calcular dividendos recebidos
        dividendos_info = calculate_total_dividendos(user_id=current_user.id)
        total_dividendos = dividendos_info.get('total_geral', 0.0)
        dividendos_por_ticker = dividendos_info.get('por_ticker', {})
        
        # Adicionar dividendos por ticker aos detalhes
        for det in detalhado:
            ticker = det['ticker']
            det['dividendos_recebidos'] = dividendos_por_ticker.get(ticker, 0.0)
        
        # PnL da carteira (posições abertas) = (preço atual - preço médio) * quantidade atual
        pnl_carteira = total_valor - total_investido
        
        # PnL total = PnL não realizado (carteira) + Dividendos recebidos + PnL realizado (vendas)
        # Fórmula: ((preço atual - preço médio) * quantidade atual) + dividendos + lucro das vendas
        pnl_total = pnl_carteira + total_dividendos + pnl_realizado
        
        # Rentabilidades
        # Rentabilidade da carteira (posições abertas) = PnL não realizado / Investido
        rentabilidade_carteira = (pnl_carteira / total_investido) if total_investido > 0 else 0.0
        
        # Rentabilidade realizada (vendas) = PnL realizado / Custo das vendas
        rentabilidade_realizada = (pnl_realizado / custo_vendas) if custo_vendas > 0 else 0.0
        
        # Investimento total = Investido em posições abertas + Custo das ações vendidas
        investimento_total = total_investido + custo_vendas
        
        # Rentabilidade total = PnL total / Investimento total
        # Onde PnL total = (preço atual - preço médio) * quantidade atual + dividendos + lucro vendas
        rentabilidade_total = (pnl_total / investimento_total) if investimento_total > 0 else 0.0
        
        return jsonify({
            'positions': detalhado,
            'total_investido': total_investido,
            'total_valor': total_valor,
            'pnl_carteira': pnl_carteira,
            'pnl_realizado': pnl_realizado,
            'total_dividendos': total_dividendos,
            'pnl_total': pnl_total,
            'custo_vendas': custo_vendas,
            'receita_vendas': receita_vendas,
            'rentabilidade_carteira': rentabilidade_carteira,
            'rentabilidade_realizada': rentabilidade_realizada,
            'rentabilidade': rentabilidade_total
        })
    except Exception as e:
        logger.error(f"❌ Erro ao calcular resumo da carteira: {str(e)}")
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/ia_recomendacoes', methods=['GET'])
@login_required
def ia_recomendacoes():
    """Gera recomendações estratégicas de IA para a carteira (versão simples)."""
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        resumo = positions_summary(user_id=current_user.id)
        pos = resumo.get('positions', [])
        
        total_investido = 0.0
        total_valor = 0.0
        detalhado = []
        
        for p in pos:
            ticker = p['ticker']
            qty = float(p['net_quantity'])
            avg_cost = float(p['avg_cost'])
            total_investido += max(qty, 0) * avg_cost
            
            preco = None
            try:
                preco_df = coletar_ultimo_pregao(ticker)
                if not preco_df.empty:
                    preco = float(preco_df['Close'].iloc[-1])
            except:
                preco = None
            
            valor_posicao = qty * (preco if preco is not None else 0.0)
            if preco is not None:
                total_valor += valor_posicao
            
            rentabilidade_simples = None
            if preco is not None and avg_cost > 0:
                rentabilidade_simples = (preco - avg_cost) / avg_cost
            
            detalhado.append({
                'ticker': ticker,
                'quantidade': qty,
                'preco_medio': avg_cost,
                'preco_ultimo': preco,
                'valor_posicao': valor_posicao if preco is not None else None,
                'rentabilidade': rentabilidade_simples
            })
        
        pnl_carteira = total_valor - total_investido
        pnl_realizado_info = calculate_realized_pnl(user_id=current_user.id)
        pnl_realizado = pnl_realizado_info.get('total_pnl_realizado', 0.0)
        pnl_total = pnl_realizado + pnl_carteira
        
        rentabilidade_carteira = (pnl_carteira / total_investido) if total_investido > 0 else 0.0
        custo_vendas = pnl_realizado_info.get('total_custo_vendas', 0.0)
        rentabilidade_realizada = (pnl_realizado / custo_vendas) if custo_vendas > 0 else 0.0
        investimento_total = total_investido + custo_vendas
        rentabilidade_total = (pnl_total / investimento_total) if investimento_total > 0 else 0.0
        receita_vendas = pnl_realizado_info.get('total_receita_vendas', 0.0)
        
        resultado_ia = analisar_carteira_com_ia(
            positions=detalhado,
            pnl_carteira=pnl_carteira,
            pnl_realizado=pnl_realizado,
            pnl_total=pnl_total,
            rentabilidade_carteira=rentabilidade_carteira,
            rentabilidade_realizada=rentabilidade_realizada,
            rentabilidade_total=rentabilidade_total,
            total_investido=total_investido,
            total_valor=total_valor,
            custo_vendas=custo_vendas,
            receita_vendas=receita_vendas
        )
        
        return jsonify(resultado_ia)
        
    except Exception as e:
        logger.error(f"❌ Erro ao gerar recomendações de IA: {str(e)}")
        traceback.print_exc()
        return jsonify({'erro': str(e), 'status': 'erro'}), 500

@app.route('/api/assessor_investimentos', methods=['GET'])
@login_required
def assessor_investimentos():
    """Análise completa da carteira com agente assessor de investimentos."""
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        logger.info(f"🤖 [ASSESSOR] Iniciando análise completa para user_id={current_user.id}")
        
        # 1. Obter posições
        resumo = positions_summary(user_id=current_user.id)
        pos = resumo.get('positions', [])
        
        if not pos:
            return jsonify({
                'status': 'erro',
                'erro': 'Nenhuma posição aberta encontrada',
                'mensagem': 'Importe operações primeiro para gerar análise completa.'
            }), 400
        
        # 2. Obter todas as operações (trades)
        trades = trades_list(user_id=current_user.id, limit=1000)
        
        # 3. Obter todos os dividendos
        dividendos = list_dividendos(user_id=current_user.id, limit=1000)
        
        # 4. Calcular métricas
        total_investido = 0.0
        total_valor = 0.0
        detalhado = []
        
        for p in pos:
            ticker = p['ticker']
            qty = float(p['net_quantity'])
            avg_cost = float(p['avg_cost'])
            total_investido += max(qty, 0) * avg_cost
            
            preco = None
            try:
                preco_df = coletar_ultimo_pregao(ticker)
                if not preco_df.empty:
                    preco = float(preco_df['Close'].iloc[-1])
            except:
                preco = None
            
            valor_posicao = qty * (preco if preco is not None else 0.0)
            if preco is not None:
                total_valor += valor_posicao
            
            rentabilidade_simples = None
            if preco is not None and avg_cost > 0:
                rentabilidade_simples = (preco - avg_cost) / avg_cost
            
            detalhado.append({
                'ticker': ticker,
                'quantidade': qty,
                'preco_medio': avg_cost,
                'preco_ultimo': preco,
                'valor_posicao': valor_posicao if preco is not None else None,
                'rentabilidade': rentabilidade_simples
            })
        
        pnl_carteira = total_valor - total_investido
        pnl_realizado_info = calculate_realized_pnl(user_id=current_user.id)
        pnl_realizado = pnl_realizado_info.get('total_pnl_realizado', 0.0)
        pnl_total = pnl_realizado + pnl_carteira
        
        rentabilidade_carteira = (pnl_carteira / total_investido) if total_investido > 0 else 0.0
        custo_vendas = pnl_realizado_info.get('total_custo_vendas', 0.0)
        rentabilidade_realizada = (pnl_realizado / custo_vendas) if custo_vendas > 0 else 0.0
        investimento_total = total_investido + custo_vendas
        rentabilidade_total = (pnl_total / investimento_total) if investimento_total > 0 else 0.0
        receita_vendas = pnl_realizado_info.get('total_receita_vendas', 0.0)
        
        # 5. Chamar assessor completo
        logger.info(f"🤖 [ASSESSOR] Chamando análise completa...")
        resultado = analisar_carteira_completa(
            user_id=current_user.id,
            positions=detalhado,
            trades=trades,
            dividendos=dividendos,
            pnl_carteira=pnl_carteira,
            pnl_realizado=pnl_realizado,
            pnl_total=pnl_total,
            rentabilidade_carteira=rentabilidade_carteira,
            rentabilidade_realizada=rentabilidade_realizada,
            rentabilidade_total=rentabilidade_total,
            total_investido=total_investido,
            total_valor=total_valor,
            custo_vendas=custo_vendas,
            receita_vendas=receita_vendas
        )
        
        logger.info(f"✅ [ASSESSOR] Análise completa concluída")
        return jsonify(resultado)
        
    except Exception as e:
        logger.error(f"❌ Erro no assessor de investimentos: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'status': 'erro',
            'erro': str(e),
            'mensagem': 'Erro ao realizar análise completa. Verifique se há API de IA configurada.'
        }), 500

@app.route('/api/template_operacoes', methods=['GET'])
@login_required
def template_operacoes():
    """Gera um modelo Excel para importação de operações."""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Operacoes"
            headers = ["date", "ticker", "side", "quantity", "price", "fees"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            exemplos = [
                ["2025-01-10", "BBSE3", "BUY", 100, 32.50, 2.50],
                ["2025-02-03", "PETR4", "BUY", 50, 39.10, 1.90],
                ["2025-02-20", "BBSE3", "SELL", 40, 34.20, 1.20]
            ]
            for row in exemplos:
                ws.append(row)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="modelo_operacoes.xlsx"
            )
        except Exception:
            csv_content = "date,ticker,side,quantity,price,fees\n" \
                          "2025-01-10,BBSE3,BUY,100,32.50,2.50\n" \
                          "2025-02-03,PETR4,BUY,50,39.10,1.90\n" \
                          "2025-02-20,BBSE3,SELL,40,34.20,1.20\n"
            resp = make_response(csv_content)
            resp.headers["Content-Type"] = "text/csv; charset=utf-8"
            resp.headers["Content-Disposition"] = "attachment; filename=modelo_operacoes.csv"
            return resp
    except Exception as e:
        logger.error(f"❌ Erro ao gerar template de operações: {e}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/trades_reset', methods=['POST'])
@login_required
def trades_reset():
    """Limpa a base de operações (todas as linhas)."""
    try:
        res = reset_trades(user_id=current_user.id)
        return jsonify(res)
    except Exception as e:
        logger.error(f"❌ Erro ao resetar trades: {e}")
        return jsonify({'erro': str(e)}), 500

# ========== ENDPOINTS DE DIVIDENDOS ==========

@app.route('/api/importar_dividendos', methods=['POST'])
@login_required
def importar_dividendos():
    """Importa dividendos via Excel (preferencial) ou CSV (fallback)."""
    try:
        if 'file' not in flask_request.files:
            return jsonify({'erro': 'Arquivo não encontrado no formulário (campo "file")'}), 400
        file = flask_request.files['file']
        if file.filename == '':
            return jsonify({'erro': 'Nome de arquivo inválido'}), 400
        filename = file.filename.lower()
        data_bytes = file.read()
        if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            from data.trades_repository import import_dividendos_excel_bytes, insert_dividendos_rows
            import pandas as pd
            from io import BytesIO
            df = pd.read_excel(BytesIO(data_bytes))
            rows = df.to_dict('records')
            resultado = insert_dividendos_rows(rows, user_id=current_user.id)
        elif filename.endswith('.csv'):
            from data.trades_repository import import_dividendos_csv_bytes, insert_dividendos_rows
            import pandas as pd
            from io import BytesIO
            df = pd.read_csv(BytesIO(data_bytes))
            rows = df.to_dict('records')
            resultado = insert_dividendos_rows(rows, user_id=current_user.id)
        else:
            from data.trades_repository import insert_dividendos_rows
            import pandas as pd
            from io import BytesIO
            if data_bytes[:2] == b'PK':
                df = pd.read_excel(BytesIO(data_bytes))
            else:
                df = pd.read_csv(BytesIO(data_bytes))
            rows = df.to_dict('records')
            resultado = insert_dividendos_rows(rows, user_id=current_user.id)
        return jsonify({'status': 'ok', 'inseridos': resultado.get('inserted', 0)})
    except Exception as e:
        logger.error(f"❌ Erro ao importar dividendos: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos', methods=['GET'])
@login_required
def listar_dividendos():
    """Lista dividendos armazenados (últimas 200)."""
    try:
        rows = list_dividendos(user_id=current_user.id)
        return jsonify({'dividendos': rows})
    except Exception as e:
        logger.error(f"❌ Erro ao listar dividendos: {e}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_recebidos/<ticker>', methods=['GET'])
@login_required
def listar_dividendos_recebidos(ticker):
    """Lista dividendos recebidos para um ticker específico."""
    try:
        from data.trades_repository import list_dividendos_por_ticker
        dividendos = list_dividendos_por_ticker(ticker, user_id=current_user.id)
        total = sum(d.get('valor_total', 0) for d in dividendos)
        return jsonify({
            'status': 'ok',
            'ticker': ticker,
            'total': len(dividendos),
            'valor_total': total,
            'dividendos': dividendos
        })
    except Exception as e:
        logger.error(f"❌ Erro ao listar dividendos recebidos para {ticker}: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/template_dividendos', methods=['GET'])
@login_required
def template_dividendos():
    """Gera um modelo Excel para importação de dividendos."""
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Dividendos"
            headers = ["data_pagamento", "ticker", "valor_por_acao", "quantidade_acoes", "tipo"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            exemplos = [
                ["2025-01-15", "BBSE3", 0.25, 100, "DIVIDENDO"],
                ["2025-02-10", "PETR4", 0.50, 50, "DIVIDENDO"],
                ["2025-03-05", "ITUB4", 0.15, 200, "JCP"]
            ]
            for row in exemplos:
                ws.append(row)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="template_dividendos.xlsx"
            )
        except ImportError:
            csv_content = "data_pagamento,ticker,valor_por_acao,quantidade_acoes,tipo\n"
            csv_content += "2025-01-15,BBSE3,0.25,100,DIVIDENDO\n"
            csv_content += "2025-02-10,PETR4,0.50,50,DIVIDENDO\n"
            csv_content += "2025-03-05,ITUB4,0.15,200,JCP\n"
            response = make_response(csv_content)
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = 'attachment; filename=template_dividendos.csv'
            return response
    except Exception as e:
        logger.error(f"❌ Erro ao gerar template de dividendos: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_reset', methods=['POST'])
@login_required
def dividendos_reset():
    """Limpa a base de dividendos (todas as linhas)."""
    try:
        res = reset_dividendos(user_id=current_user.id)
        return jsonify(res)
    except Exception as e:
        logger.error(f"❌ Erro ao resetar dividendos: {e}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_buscar_automatico', methods=['POST'])
@login_required
def buscar_dividendos_automatico():
    """
    Busca dividendos automaticamente da API Brapi.dev para as ações da carteira.
    Parâmetros opcionais (JSON):
    - tickers: Lista de tickers específicos (se não fornecido, usa posições abertas)
    - forcar: Se True, força atualização mesmo se dados são recentes
    """
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        data = flask_request.get_json() or {}
        tickers = data.get('tickers', [])
        forcar = data.get('forcar', False)
        
        logger.info(f"🔄 [API] Sincronizando dividendos automaticamente...")
        logger.info(f"   Tickers: {tickers if tickers else 'Todas as posições abertas'}")
        logger.info(f"   Forçar atualização: {forcar}")
        
        resultado = sincronizar_dividendos_automatico(tickers=tickers if tickers else None, forcar_atualizacao=forcar, user_id=current_user.id)
        
        if resultado.get('status') == 'erro':
            return jsonify(resultado), 400
        
        logger.info(f"✅ [API] Busca concluída: {resultado.get('total_importados', 0)} dividendos importados")
        
        total_importados = resultado.get('total_importados', 0)
        total_em_cache = resultado.get('total_em_cache', 0)
        
        if total_importados > 0:
            mensagem = f"Sincronização concluída! {total_importados} novos dividendos importados."
        elif total_em_cache > 0:
            mensagem = f"Dados já estão atualizados! {total_em_cache} tickers com dados em cache recentes."
        else:
            mensagem = "Sincronização concluída. Nenhum novo dividendo encontrado."
        
        return jsonify({
            'status': 'ok',
            'mensagem': mensagem,
            'total_encontrados': resultado.get('total_encontrados', 0),
            'total_importados': total_importados,
            'total_em_cache': total_em_cache,
            'tickers_processados': resultado.get('tickers_processados', 0),
            'erros': resultado.get('erros', [])
        })
    except Exception as e:
        logger.error(f"❌ Erro ao buscar dividendos automaticamente: {str(e)}")
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_ticker/<ticker>', methods=['GET'])
@login_required
def buscar_dividendos_ticker(ticker):
    """Busca dividendos de um ticker específico via API (sem importar no banco)."""
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        logger.info(f"🔍 [API] Buscando dividendos para {ticker}...")
        dividendos = coletar_dividendos_brapi(ticker, limit=100)
        
        return jsonify({
            'status': 'ok',
            'ticker': ticker,
            'total': len(dividendos),
            'dividendos': dividendos
        })
    except Exception as e:
        logger.error(f"❌ Erro ao buscar dividendos para {ticker}: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_limpar_invalidos', methods=['POST'])
@login_required
def limpar_dividendos_invalidos():
    """
    Executa o agente de limpeza para remover dividendos inválidos.
    Remove dividendos onde data_ex_dividendo < primeira_data_compra.
    """
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        from data.trades_repository import limpar_dividendos_invalidos
        
        logger.info(f"🧹 [API] Executando limpeza de dividendos inválidos...")
        resultado = limpar_dividendos_invalidos(user_id=current_user.id)
        
        removidos = resultado.get('total_removidos', 0)
        verificados = resultado.get('total_verificados', 0)
        removidos_por_ticker = resultado.get('removidos_por_ticker', {})
        
        if removidos > 0:
            mensagem = f"Limpeza concluída! {removidos} dividendos inválidos removidos."
        else:
            mensagem = "Limpeza concluída! Nenhum dividendo inválido encontrado."
        
        return jsonify({
            'status': 'ok',
            'mensagem': mensagem,
            'total_verificados': verificados,
            'total_removidos': removidos,
            'removidos_por_ticker': removidos_por_ticker
        })
    except Exception as e:
        logger.error(f"❌ Erro ao limpar dividendos inválidos: {str(e)}")
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/tickers')
def get_tickers():
    """Retorna lista de tickers disponíveis"""
    try:
        tickers_lista = [
            'BBSE3.SA', 'CMIG4.SA', 'CSMG3.SA', 'ITUB4.SA', 
            'PETR4.SA', 'SANB11.SA', 'SYN3.SA'
        ]
        
        if MODULOS_CARREGADOS and 'TICKERS' in globals():
            return jsonify({'tickers': TICKERS})
        else:
            return jsonify({'tickers': tickers_lista})
    except Exception as e:
        return jsonify({'tickers': tickers_lista, 'erro': str(e)})

# Handlers de erro
@app.errorhandler(404)
def not_found(e):
    return jsonify({'erro': 'Endpoint não encontrado'}), 404

@app.errorhandler(500)
def internal_error(e):
    logger.error(f"❌ Erro 500: {str(e)}")
    traceback.print_exc()
    return jsonify({'erro': 'Erro interno do servidor'}), 500

@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"❌ Erro não tratado: {str(e)}")
    traceback.print_exc()
    return jsonify({'erro': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    
    try:
        print("="*70)
        print("🚀 Dashboard Algoritimo Trade - DADOS REAIS DO MERCADO")
        print("="*70)
        print(f"📊 Acesse: http://localhost:{port}")
        print("="*70)
        
        if MODULOS_CARREGADOS:
            print("✅ Todos os módulos carregados - Dados REAIS ativos")
        else:
            print("⚠️  Alguns módulos não carregados - Verifique dependências")
        
        print("\nPressione Ctrl+C para parar\n")
        
        app.run(debug=debug_mode, host='0.0.0.0', port=port, use_reloader=False)
    except KeyboardInterrupt:
        print("\n\n⚠️  Servidor interrompido pelo usuário")
    except Exception as e:
        print(f"\n❌ Erro ao iniciar servidor: {e}")
        traceback.print_exc()

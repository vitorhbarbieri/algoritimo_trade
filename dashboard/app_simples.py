"""
Dashboard com dados reais do mercado - sem mockados
"""
from flask import Flask, render_template, jsonify, request as flask_request, send_file, make_response
import io
import sys
import os
from datetime import datetime
import pandas as pd
import logging

# Adicionar diretórios ao path
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, base_dir)

# Carregar variáveis de ambiente do arquivo .env (se existir)
# Procura o .env na raiz do projeto (base_dir)
try:
    from dotenv import load_dotenv
    env_path = os.path.join(base_dir, '.env')
    load_dotenv(env_path)
except ImportError:
    # python-dotenv não instalado, continuar sem ele
    pass

# Configurar logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Importar módulos reais
try:
    from data.price_collector import coletar_precos, coletar_ultimo_pregao
    from data.trades_repository import (
        init_db as trades_init_db, import_csv as trades_import_csv, list_trades as trades_list, 
        positions_summary, reset_trades, calculate_realized_pnl,
        import_dividendos_csv_bytes, import_dividendos_excel_bytes, list_dividendos, 
        calculate_total_dividendos, reset_dividendos
    )
    from data.dividendos_collector import sincronizar_dividendos_automatico, coletar_dividendos_brapi
    from data.trades_repository import verificar_necessidade_sincronizacao_dividendos
    from data.news_collector import coletar_noticias_brasileiras
    from features.technical_indicators import calcular_todos_indicadores
    from features.statistical_features import calcular_todas_features_estatisticas
    from features.sentiment_engine import analisar_sentimento_noticias
    from strategies.trend_strategy import gerar_sinal_tendencia
    from strategies.mean_reversion_strategy import gerar_sinal_reversao
    from strategies.news_strategy import gerar_sinal_noticias
    from core.signal_orchestrator import SignalOrchestrator
    from core.trade_executor import TradeExecutor
    from core.risk_manager import RiskManager
    from core.ia_advisor import analisar_carteira_com_ia
    from utils.config import TICKERS, PESOS_ESTRATEGIAS, CAPITAL_INICIAL
    MODULOS_CARREGADOS = True
except ImportError as e:
    print(f"⚠️  Aviso: Alguns módulos não foram carregados: {e}")
    MODULOS_CARREGADOS = False
    TICKERS = ['BBSE3.SA', 'CMIG4.SA', 'CSMG3.SA', 'ITUB4.SA', 'PETR4.SA', 'SANB11.SA', 'SYN3.SA']
    PESOS_ESTRATEGIAS = {'trend': 0.4, 'reversao': 0.3, 'news': 0.3}
    CAPITAL_INICIAL = 10000.0

app = Flask(__name__)

# Estado do sistema
try:
    executor = TradeExecutor(modo_mock=True) if MODULOS_CARREGADOS else None
    # Inicializar base de dados de trades
    try:
        trades_init_db()
        print("🗄️  Base de dados de operações inicializada.")
    except Exception as e:
        print(f"⚠️  Não foi possível inicializar DB de operações: {e}")
except:
    executor = None

capital_atual = CAPITAL_INICIAL

@app.route('/')
def index():
    """Home - Importar operações e visualizar carteira"""
    try:
        return render_template('home.html')
    except Exception as e:
        return f"<h1>Home</h1><p>Erro ao carregar template: {e}</p>", 500

@app.route('/api/status')
def get_status():
    """Status do sistema"""
    try:
        capital_inicial = CAPITAL_INICIAL if 'CAPITAL_INICIAL' in globals() else 10000.0
        retorno = ((capital_atual - capital_inicial) / capital_inicial * 100) if capital_inicial > 0 else 0
        
        posicoes_abertas = 0
        total_operacoes = 0
        
        if executor is not None:
            try:
                posicoes = executor.obter_posicoes_abertas()
                posicoes_abertas = len(posicoes) if posicoes else 0
                
                historico = executor.obter_historico()
                total_operacoes = len(historico) if not historico.empty else 0
            except:
                pass
        
        return jsonify({
            'capital_atual': capital_atual,
            'capital_inicial': capital_inicial,
            'retorno_percentual': retorno,
            'posicoes_abertas': posicoes_abertas,
            'total_operacoes': total_operacoes,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/analisar/<ticker>')
def analisar_ticker(ticker):
    """Análise REAL com dados do mercado"""
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados. Verifique as dependências.'}), 500
        
        # 1. Coletar dados REAIS de preços
        print(f"\n{'='*70}")
        print(f"🔍 [DASHBOARD] Iniciando análise para {ticker}")
        print(f"{'='*70}")
        print(f"📡 [DASHBOARD] Coletando dados reais de preços para {ticker}...")
        
        try:
            # 1) Buscar último pregão (modo leve para evitar 429)
            df_ultimo = coletar_ultimo_pregao(ticker)
            print(f"✅ [DASHBOARD] Último pregão coletado: {len(df_ultimo)} período")
            
            # 2) Tentar expandir para um período maior; se falhar, seguir em modo mínimo
            modo_minimo = False
            try:
                df_precos = coletar_precos(ticker, periodo='1mo', intervalo='1d')
                print(f"✅ [DASHBOARD] Período expandido coletado: {len(df_precos)} períodos")
            except Exception as e_expand:
                print(f"⚠️  [DASHBOARD] Não foi possível coletar período de 1 mês (seguindo com último pregão): {e_expand}")
                df_precos = df_ultimo.copy()
                modo_minimo = True
        except Exception as e:
            print(f"❌ [DASHBOARD] Erro ao coletar preços: {e}")
            return jsonify({
                'erro': f'Erro ao coletar preços de {ticker}',
                'detalhe': str(e),
                'sugestao': 'Verifique se o ticker está correto e se há dados disponíveis no yfinance'
            }), 404
        
        if df_precos.empty:
            print(f"❌ [DASHBOARD] DataFrame de preços está vazio para {ticker}")
            return jsonify({
                'erro': f'Nenhum dado encontrado para {ticker}',
                'sugestao': 'O ticker pode estar incorreto ou não ter dados disponíveis no período solicitado'
            }), 404
        
        # 2. Coletar notícias REAIS
        print(f"📰 [DASHBOARD] Coletando notícias...")
        noticias = []
        try:
            noticias = coletar_noticias_brasileiras()
            print(f"✅ [DASHBOARD] Notícias coletadas: {len(noticias)}")
        except Exception as e:
            print(f"⚠️  [DASHBOARD] Erro ao coletar notícias: {e}")
        
        # 3. Calcular indicadores REAIS
        df = df_precos.copy()
        if len(df_precos) > 1:
            print(f"📊 [DASHBOARD] Calculando indicadores técnicos...")
            try:
                df = calcular_todos_indicadores(df_precos)
                print(f"✅ [DASHBOARD] Indicadores técnicos calculados")
                print(f"   Colunas disponíveis: {list(df.columns)[:10]}...")
            except Exception as e:
                print(f"❌ [DASHBOARD] Erro ao calcular indicadores técnicos: {e}")
                import traceback
                traceback.print_exc()
                # Em modo mínimo, seguimos sem indicadores
        else:
            print(f"ℹ️  [DASHBOARD] Apenas 1 período disponível. Pulando indicadores (modo mínimo).")
        
        print(f"📈 [DASHBOARD] Calculando features estatísticas...")
        try:
            if len(df) > 1:
                df = calcular_todas_features_estatisticas(df)
                print(f"✅ [DASHBOARD] Features estatísticas calculadas")
            else:
                print(f"ℹ️  [DASHBOARD] Pulando features estatísticas (modo mínimo).")
        except Exception as e:
            print(f"⚠️  [DASHBOARD] Erro ao calcular features estatísticas: {e}")
        
        # 4. Gerar sinais REAIS
        print(f"🎯 [DASHBOARD] Gerando sinais das estratégias...")
        sinal_trend = pd.Series(0, index=df.index)
        sinal_reversao = pd.Series(0, index=df.index)
        sinal_news_valor = 0
        sentimento = 0.0
        
        try:
            if len(df) > 1:
                print(f"   📈 Gerando sinal de tendência...")
                sinal_trend = gerar_sinal_tendencia(df)
                sinais_trend_count = (sinal_trend != 0).sum()
                print(f"   ✅ Sinal de tendência: {sinais_trend_count} sinais não-neutros")
            else:
                print(f"   ℹ️  Pulando sinal de tendência (modo mínimo)")
        except Exception as e:
            print(f"   ❌ Erro na estratégia de tendência: {e}")
            import traceback
            traceback.print_exc()
        
        try:
            if len(df) > 1:
                print(f"   🔄 Gerando sinal de reversão...")
                sinal_reversao = gerar_sinal_reversao(df)
                sinais_reversao_count = (sinal_reversao != 0).sum()
                print(f"   ✅ Sinal de reversão: {sinais_reversao_count} sinais não-neutros")
            else:
                print(f"   ℹ️  Pulando sinal de reversão (modo mínimo)")
        except Exception as e:
            print(f"   ❌ Erro na estratégia de reversão: {e}")
            import traceback
            traceback.print_exc()
        
        try:
            print(f"   📰 Analisando sentimento das notícias...")
            sentimento = analisar_sentimento_noticias(noticias)
            sinal_news_valor = gerar_sinal_noticias(noticias)
            print(f"   ✅ Sentimento: {sentimento:+.3f}, Sinal: {sinal_news_valor}")
        except Exception as e:
            print(f"   ❌ Erro na estratégia de notícias: {e}")
            import traceback
            traceback.print_exc()
        
        sinal_news = pd.Series([sinal_news_valor] * len(df), index=df.index)
        
        # 5. Combinar sinais REAIS
        print(f"🎼 [DASHBOARD] Combinando sinais (orquestração)...")
        try:
            orchestrator = SignalOrchestrator(pesos=PESOS_ESTRATEGIAS)
            sinal_final = orchestrator.combinar_sinais(sinal_trend, sinal_reversao, sinal_news)
            confianca = orchestrator.calcular_confianca(sinal_trend, sinal_reversao, sinal_news)
            print(f"   ✅ Sinais combinados: {len(sinal_final)} períodos processados")
        except Exception as e:
            print(f"   ❌ Erro na orquestração: {e}")
            import traceback
            traceback.print_exc()
            sinal_final = pd.Series(0, index=df.index)
            confianca = pd.Series(0.5, index=df.index)
        
        # 6. Dados REAIS do último período
        print(f"📊 [DASHBOARD] Extraindo dados do último período...")
        ultimo_preco = float(df['Close'].iloc[-1])
        rsi = float(df['RSI'].iloc[-1]) if 'RSI' in df.columns and pd.notna(df['RSI'].iloc[-1]) else None
        macd = float(df['MACD'].iloc[-1]) if 'MACD' in df.columns and pd.notna(df['MACD'].iloc[-1]) else None
        ultimo_sinal = int(sinal_final.iloc[-1]) if len(sinal_final) > 0 else 0
        ultima_confianca = float(confianca.iloc[-1]) if len(confianca) > 0 else 0.5
        
        print(f"   Preço atual: R$ {ultimo_preco:.2f}")
        print(f"   RSI: {rsi:.2f}" if rsi else "   RSI: N/A")
        print(f"   MACD: {macd:.4f}" if macd else "   MACD: N/A")
        print(f"   Sinal final: {ultimo_sinal} (Confiança: {ultima_confianca:.1%})")
        
        # 7. Variação do preço (último vs anterior)
        if len(df) >= 2:
            preco_anterior = float(df['Close'].iloc[-2])
            variacao_percentual = ((ultimo_preco - preco_anterior) / preco_anterior) * 100
            print(f"   Variação: {variacao_percentual:+.2f}% (de R$ {preco_anterior:.2f} para R$ {ultimo_preco:.2f})")
        else:
            variacao_percentual = 0.0
            print(f"   Variação: N/A (apenas 1 período disponível)")
        
        print(f"{'='*70}")
        print(f"✅ [DASHBOARD] Análise concluída para {ticker}")
        print(f"{'='*70}\n")
        
        return jsonify({
            'ticker': ticker,
            'preco_atual': ultimo_preco,
            'variacao_percentual': variacao_percentual,
            'sinal_final': ultimo_sinal,
            'confianca': ultima_confianca,
            'sentimento_noticias': float(sentimento),
            'rsi': rsi,
            'macd': macd,
            'sinal_trend': int(sinal_trend.iloc[-1]) if len(sinal_trend) > 0 else 0,
            'sinal_reversao': int(sinal_reversao.iloc[-1]) if len(sinal_reversao) > 0 else 0,
            'sinal_news': int(sinal_news_valor),
            'timestamp': datetime.now().isoformat(),
            'fonte_dados': 'MERCADO REAL',
            'modo_minimo': True if 'modo_minimo' in locals() and modo_minimo else False
        })
    
    except Exception as e:
        import traceback
        error_msg = f"{str(e)}\n{traceback.format_exc()}"
        print(f"Erro completo: {error_msg}")
        return jsonify({'erro': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/operacoes')
def get_operacoes():
    """Histórico de operações"""
    try:
        if executor is None:
            return jsonify({'operacoes': []})
        
        historico = executor.obter_historico()
        
        if historico.empty:
            return jsonify({'operacoes': []})
        
        operacoes = historico.to_dict('records')
        for op in operacoes:
            if 'timestamp' in op and pd.notna(op['timestamp']):
                op['timestamp'] = op['timestamp'].isoformat() if hasattr(op['timestamp'], 'isoformat') else str(op['timestamp'])
        
        return jsonify({'operacoes': operacoes[-10:]})
    except Exception as e:
        return jsonify({'operacoes': [], 'erro': str(e)})

@app.route('/api/importar_operacoes', methods=['POST'])
def importar_operacoes():
    """Importa operações via Excel (preferencial) ou CSV (fallback).
    Campos esperados: date, ticker, side (BUY/SELL), quantity, price, fees
    """
    try:
        if 'file' not in flask_request.files:
            return jsonify({'erro': 'Arquivo não encontrado no formulário (campo "file")'}), 400
        file = flask_request.files['file']
        if file.filename == '':
            return jsonify({'erro': 'Nome de arquivo inválido'}), 400
        filename = file.filename.lower()
        data_bytes = file.read()
        # Preferir Excel
        if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            from data.trades_repository import import_excel_bytes
            resultado = import_excel_bytes(data_bytes)
        elif filename.endswith('.csv'):
            from data.trades_repository import import_csv_bytes
            resultado = import_csv_bytes(data_bytes)
        else:
            # Tentar detectar pelo conteúdo: se começar com PK é zip/xlsx
            if data_bytes[:2] == b'PK':
                from data.trades_repository import import_excel_bytes
                resultado = import_excel_bytes(data_bytes)
            else:
                from data.trades_repository import import_csv_bytes
                resultado = import_csv_bytes(data_bytes)
        return jsonify({'status': 'ok', 'inseridos': resultado.get('inserted', 0)})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/trades', methods=['GET'])
def listar_trades():
    """Lista operações armazenadas (últimas 200)."""
    try:
        rows = trades_list()
        return jsonify({'trades': rows})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/portfolio_resumo', methods=['GET'])
def portfolio_resumo():
    """Consolida posições, custos médios e avalia pelo último pregão."""
    try:
        resumo = positions_summary()
        pos = resumo.get('positions', [])
        total_investido = 0.0
        total_valor = 0.0
        detalhado = []
        hoje = datetime.now().date()
        
        for p in pos:
            ticker = p['ticker']
            qty = float(p['net_quantity'])
            avg_cost = float(p['avg_cost'])
            first_buy_date_str = p.get('first_buy_date')
            total_investido += max(qty, 0) * avg_cost
            
            # Tentar buscar preço, mas continuar mesmo se falhar
            preco = None
            preco_disponivel = False
            try:
                preco_df = coletar_ultimo_pregao(ticker)
                if not preco_df.empty:
                    preco = float(preco_df['Close'].iloc[-1])
                    preco_disponivel = True
            except Exception as e_preco:
                logger.warning(f"⚠️  Não foi possível obter preço para {ticker}: {str(e_preco)[:100]}")
                preco = None
                preco_disponivel = False
            
            # Calcular valor da posição (0 se preço não disponível)
            valor_posicao = qty * (preco if preco is not None else 0.0)
            if preco is not None:
                total_valor += valor_posicao
            
            # Calcular rentabilidade simples e anualizada
            rentabilidade_simples = None
            rentabilidade_anualizada = None
            if preco is not None and avg_cost > 0:
                try:
                    # Calcular rentabilidade simples: (preco_atual - preco_medio) / preco_medio
                    rentabilidade_simples = (preco - avg_cost) / avg_cost
                    
                    # Calcular rentabilidade anualizada se tiver data da primeira compra
                    if first_buy_date_str:
                        try:
                            # Calcular dias desde a primeira compra
                            first_buy_date = datetime.strptime(first_buy_date_str, "%Y-%m-%d").date()
                            dias = (hoje - first_buy_date).days
                            
                            # Anualizar: (1 + r)^(365/dias) - 1
                            if dias > 0:
                                rentabilidade_anualizada = ((1 + rentabilidade_simples) ** (365.0 / dias)) - 1
                            elif dias == 0:
                                # Se comprou hoje, usar rentabilidade simples
                                rentabilidade_anualizada = rentabilidade_simples
                        except Exception as e_rent:
                            logger.warning(f"⚠️  Erro ao calcular rentabilidade anualizada para {ticker}: {str(e_rent)[:100]}")
                            rentabilidade_anualizada = None
                except Exception as e_rent_simples:
                    logger.warning(f"⚠️  Erro ao calcular rentabilidade simples para {ticker}: {str(e_rent_simples)[:100]}")
                    rentabilidade_simples = None
            
            detalhado.append({
                'ticker': ticker,
                'quantidade': qty,
                'preco_medio': avg_cost,
                'preco_ultimo': preco if preco is not None else None,
                'preco_disponivel': preco_disponivel,
                'valor_posicao': valor_posicao if preco is not None else None,
                'rentabilidade': rentabilidade_simples,
                'rentabilidade_anualizada': rentabilidade_anualizada,
                'first_buy_date': first_buy_date_str
            })
        
        # Calcular PnL realizado (vendas executadas)
        pnl_realizado_info = calculate_realized_pnl()
        pnl_realizado = pnl_realizado_info.get('total_pnl_realizado', 0.0)
        custo_vendas = pnl_realizado_info.get('total_custo_vendas', 0.0)
        receita_vendas = pnl_realizado_info.get('total_receita_vendas', 0.0)
        
        # Sincronizar dividendos automaticamente (em background, não bloqueia resposta)
        # Verifica se precisa sincronizar para cada ticker e sincroniza se necessário
        try:
            tickers_para_sincronizar = [p['ticker'] for p in pos if verificar_necessidade_sincronizacao_dividendos(p['ticker'])]
            if tickers_para_sincronizar:
                logger.info(f"🔄 [PORTFOLIO] Sincronizando dividendos para {len(tickers_para_sincronizar)} tickers em background...")
                # Executar em thread para não bloquear a resposta
                import threading
                def sync_bg():
                    try:
                        sincronizar_dividendos_automatico(tickers=tickers_para_sincronizar, forcar_atualizacao=False)
                    except Exception as e:
                        logger.warning(f"⚠️  [PORTFOLIO] Erro ao sincronizar dividendos em background: {str(e)[:100]}")
                threading.Thread(target=sync_bg, daemon=True).start()
        except Exception as e:
            logger.warning(f"⚠️  [PORTFOLIO] Erro ao verificar sincronização de dividendos: {str(e)[:100]}")
        
        # Calcular dividendos recebidos (usa dados do banco, que podem estar sendo atualizados em background)
        dividendos_info = calculate_total_dividendos()
        total_dividendos = dividendos_info.get('total_geral', 0.0)
        dividendos_por_ticker = dividendos_info.get('por_ticker', {})
        
        # Adicionar dividendos por ticker aos detalhes
        for det in detalhado:
            ticker = det['ticker']
            det['dividendos_recebidos'] = dividendos_por_ticker.get(ticker, 0.0)
        
        # PnL da carteira (posições abertas)
        pnl_carteira = total_valor - total_investido
        
        # PnL total (realizado + carteira + dividendos)
        pnl_total = pnl_realizado + pnl_carteira + total_dividendos
        
        # Rentabilidade da carteira (posições abertas) - SEM dividendos
        rentabilidade_carteira = (pnl_carteira / total_investido) if total_investido > 0 else 0.0
        
        # Rentabilidade realizada (vendas) - SEM dividendos
        rentabilidade_realizada = (pnl_realizado / custo_vendas) if custo_vendas > 0 else 0.0
        
        # Rentabilidade total (considerando investimento total: carteira + custo das vendas + dividendos)
        investimento_total = total_investido + custo_vendas
        # Rentabilidade total = (PnL total + dividendos) / investimento total
        rentabilidade_total = (pnl_total / investimento_total) if investimento_total > 0 else 0.0
        
        return jsonify({
            'positions': detalhado,
            'total_investido': total_investido,  # Investido em posições abertas
            'total_valor': total_valor,  # Valor atual das posições abertas
            'pnl_carteira': pnl_carteira,  # PnL não realizado (posições abertas)
            'pnl_realizado': pnl_realizado,  # PnL realizado (vendas executadas)
            'total_dividendos': total_dividendos,  # Total de dividendos recebidos
            'pnl_total': pnl_total,  # PnL total (realizado + carteira + dividendos)
            'custo_vendas': custo_vendas,  # Custo das ações vendidas
            'receita_vendas': receita_vendas,  # Receita das vendas
            'rentabilidade_carteira': rentabilidade_carteira,  # Rentabilidade das posições abertas (sem dividendos)
            'rentabilidade_realizada': rentabilidade_realizada,  # Rentabilidade das vendas (sem dividendos)
            'rentabilidade': rentabilidade_total  # Rentabilidade total (incluindo dividendos)
        })
    except Exception as e:
        logger.error(f"❌ Erro ao calcular resumo da carteira: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/ia_recomendacoes', methods=['GET'])
def ia_recomendacoes():
    """Gera recomendações estratégicas de IA para a carteira."""
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        # Buscar dados da carteira
        resumo = positions_summary()
        pos = resumo.get('positions', [])
        
        # Calcular métricas básicas
        total_investido = 0.0
        total_valor = 0.0
        detalhado = []
        hoje = datetime.now().date()
        
        for p in pos:
            ticker = p['ticker']
            qty = float(p['net_quantity'])
            avg_cost = float(p['avg_cost'])
            total_investido += max(qty, 0) * avg_cost
            
            # Tentar buscar preço
            preco = None
            try:
                preco_df = coletar_ultimo_pregao(ticker)
                if not preco_df.empty:
                    preco = float(preco_df['Close'].iloc[-1])
            except:
                preco = None
            
            valor_posicao = qty * (preco if preco is not None else 0.0)
            if preco is not None:
                total_valor += valor_posicao
            
            # Calcular rentabilidade simples
            rentabilidade_simples = None
            if preco is not None and avg_cost > 0:
                rentabilidade_simples = (preco - avg_cost) / avg_cost
            
            detalhado.append({
                'ticker': ticker,
                'quantidade': qty,
                'preco_medio': avg_cost,
                'preco_ultimo': preco,
                'valor_posicao': valor_posicao if preco is not None else None,
                'rentabilidade': rentabilidade_simples
            })
        
        # Calcular PnL
        pnl_carteira = total_valor - total_investido
        pnl_realizado_info = calculate_realized_pnl()
        pnl_realizado = pnl_realizado_info.get('total_pnl_realizado', 0.0)
        pnl_total = pnl_realizado + pnl_carteira
        
        rentabilidade_carteira = (pnl_carteira / total_investido) if total_investido > 0 else 0.0
        custo_vendas = pnl_realizado_info.get('total_custo_vendas', 0.0)
        rentabilidade_realizada = (pnl_realizado / custo_vendas) if custo_vendas > 0 else 0.0
        investimento_total = total_investido + custo_vendas
        rentabilidade_total = (pnl_total / investimento_total) if investimento_total > 0 else 0.0
        receita_vendas = pnl_realizado_info.get('total_receita_vendas', 0.0)
        
        # Chamar IA para análise
        resultado_ia = analisar_carteira_com_ia(
            positions=detalhado,
            pnl_carteira=pnl_carteira,
            pnl_realizado=pnl_realizado,
            pnl_total=pnl_total,
            rentabilidade_carteira=rentabilidade_carteira,
            rentabilidade_realizada=rentabilidade_realizada,
            rentabilidade_total=rentabilidade_total,
            total_investido=total_investido,
            total_valor=total_valor,
            custo_vendas=custo_vendas,
            receita_vendas=receita_vendas
        )
        
        return jsonify(resultado_ia)
        
    except Exception as e:
        logger.error(f"❌ Erro ao gerar recomendações de IA: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e), 'status': 'erro'}), 500

@app.route('/api/template_operacoes', methods=['GET'])
def template_operacoes():
    """
    Gera um modelo Excel para importação de operações.
    Colunas: date, ticker, side (BUY/SELL), quantity, price, fees
    """
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Operacoes"
            headers = ["date", "ticker", "side", "quantity", "price", "fees"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Exemplos
            exemplos = [
                ["2025-01-10", "BBSE3", "BUY", 100, 32.50, 2.50],
                ["2025-02-03", "PETR4", "BUY", 50, 39.10, 1.90],
                ["2025-02-20", "BBSE3", "SELL", 40, 34.20, 1.20]
            ]
            for row in exemplos:
                ws.append(row)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="modelo_operacoes.xlsx"
            )
        except Exception:
            # Fallback CSV
            csv_content = "date,ticker,side,quantity,price,fees\n" \
                          "2025-01-10,BBSE3,BUY,100,32.50,2.50\n" \
                          "2025-02-03,PETR4,BUY,50,39.10,1.90\n" \
                          "2025-02-20,BBSE3,SELL,40,34.20,1.20\n"
            resp = make_response(csv_content)
            resp.headers["Content-Type"] = "text/csv; charset=utf-8"
            resp.headers["Content-Disposition"] = "attachment; filename=modelo_operacoes.csv"
            return resp
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/trades_reset', methods=['POST'])
def trades_reset():
    """Limpa a base de operações (todas as linhas)."""
    try:
        res = reset_trades()
        return jsonify(res)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

# ========== ENDPOINTS DE DIVIDENDOS ==========

@app.route('/api/importar_dividendos', methods=['POST'])
def importar_dividendos():
    """Importa dividendos via Excel (preferencial) ou CSV (fallback).
    Campos esperados: data_pagamento, ticker, valor_por_acao, quantidade_acoes, tipo (opcional: DIVIDENDO/JCP/RENDIMENTO)
    """
    try:
        if 'file' not in flask_request.files:
            return jsonify({'erro': 'Arquivo não encontrado no formulário (campo "file")'}), 400
        file = flask_request.files['file']
        if file.filename == '':
            return jsonify({'erro': 'Nome de arquivo inválido'}), 400
        filename = file.filename.lower()
        data_bytes = file.read()
        # Preferir Excel
        if filename.endswith('.xlsx') or filename.endswith('.xlsm'):
            resultado = import_dividendos_excel_bytes(data_bytes)
        elif filename.endswith('.csv'):
            resultado = import_dividendos_csv_bytes(data_bytes)
        else:
            # Tentar detectar pelo conteúdo: se começar com PK é zip/xlsx
            if data_bytes[:2] == b'PK':
                resultado = import_dividendos_excel_bytes(data_bytes)
            else:
                resultado = import_dividendos_csv_bytes(data_bytes)
        return jsonify({'status': 'ok', 'inseridos': resultado.get('inserted', 0)})
    except Exception as e:
        logger.error(f"❌ Erro ao importar dividendos: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos', methods=['GET'])
def listar_dividendos():
    """Lista dividendos armazenados (últimas 200)."""
    try:
        rows = list_dividendos()
        return jsonify({'dividendos': rows})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_recebidos/<ticker>', methods=['GET'])
def listar_dividendos_recebidos(ticker):
    """Lista dividendos recebidos para um ticker específico."""
    try:
        from data.trades_repository import list_dividendos_por_ticker
        dividendos = list_dividendos_por_ticker(ticker)
        
        # Calcular total
        total = sum(d.get('valor_total', 0) for d in dividendos)
        
        return jsonify({
            'status': 'ok',
            'ticker': ticker,
            'total': len(dividendos),
            'valor_total': total,
            'dividendos': dividendos
        })
    except Exception as e:
        logger.error(f"❌ Erro ao listar dividendos recebidos para {ticker}: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/template_dividendos', methods=['GET'])
def template_dividendos():
    """
    Gera um modelo Excel para importação de dividendos.
    Colunas: data_pagamento, ticker, valor_por_acao, quantidade_acoes, tipo
    """
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Dividendos"
            headers = ["data_pagamento", "ticker", "valor_por_acao", "quantidade_acoes", "tipo"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Exemplos
            exemplos = [
                ["2025-01-15", "BBSE3", 0.25, 100, "DIVIDENDO"],
                ["2025-02-10", "PETR4", 0.50, 50, "DIVIDENDO"],
                ["2025-03-05", "ITUB4", 0.15, 200, "JCP"]
            ]
            for row in exemplos:
                ws.append(row)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            return send_file(
                bio,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="template_dividendos.xlsx"
            )
        except ImportError:
            # Fallback CSV
            csv_content = "data_pagamento,ticker,valor_por_acao,quantidade_acoes,tipo\n"
            csv_content += "2025-01-15,BBSE3,0.25,100,DIVIDENDO\n"
            csv_content += "2025-02-10,PETR4,0.50,50,DIVIDENDO\n"
            csv_content += "2025-03-05,ITUB4,0.15,200,JCP\n"
            response = make_response(csv_content)
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = 'attachment; filename=template_dividendos.csv'
            return response
    except Exception as e:
        logger.error(f"❌ Erro ao gerar template de dividendos: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_reset', methods=['POST'])
def dividendos_reset():
    """Limpa a base de dividendos (todas as linhas)."""
    try:
        res = reset_dividendos()
        return jsonify(res)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_buscar_automatico', methods=['POST'])
def buscar_dividendos_automatico():
    """
    Busca dividendos automaticamente da API Brapi.dev para as ações da carteira.
    Parâmetros opcionais (JSON):
    - tickers: Lista de tickers específicos (se não fornecido, usa posições abertas)
    - data_inicio: Data de início para filtrar (formato: 'YYYY-MM-DD')
    """
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        # Obter parâmetros do request
        data = flask_request.get_json() or {}
        tickers = data.get('tickers', [])
        data_inicio = data.get('data_inicio', None)
        
        forcar = data.get('forcar', False)
        logger.info(f"🔄 [API] Sincronizando dividendos automaticamente...")
        logger.info(f"   Tickers: {tickers if tickers else 'Todas as posições abertas'}")
        logger.info(f"   Forçar atualização: {forcar}")
        
        # Sincronizar dividendos (usa cache inteligente)
        resultado = sincronizar_dividendos_automatico(tickers=tickers if tickers else None, forcar_atualizacao=forcar)
        
        if resultado.get('status') == 'erro':
            return jsonify(resultado), 400
        
        logger.info(f"✅ [API] Busca concluída: {resultado.get('total_importados', 0)} dividendos importados")
        
        total_importados = resultado.get('total_importados', 0)
        total_em_cache = resultado.get('total_em_cache', 0)
        
        if total_importados > 0:
            mensagem = f"Sincronização concluída! {total_importados} novos dividendos importados."
        elif total_em_cache > 0:
            mensagem = f"Dados já estão atualizados! {total_em_cache} tickers com dados em cache recentes."
        else:
            mensagem = "Sincronização concluída. Nenhum novo dividendo encontrado."
        
        return jsonify({
            'status': 'ok',
            'mensagem': mensagem,
            'total_encontrados': resultado.get('total_encontrados', 0),
            'total_importados': total_importados,
            'total_em_cache': total_em_cache,
            'tickers_processados': resultado.get('tickers_processados', 0),
            'erros': resultado.get('erros', [])
        })
    except Exception as e:
        logger.error(f"❌ Erro ao buscar dividendos automaticamente: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/dividendos_ticker/<ticker>', methods=['GET'])
def buscar_dividendos_ticker(ticker):
    """
    Busca dividendos de um ticker específico via API (sem importar no banco).
    Útil para visualizar dividendos antes de importar.
    """
    try:
        if not MODULOS_CARREGADOS:
            return jsonify({'erro': 'Módulos não carregados'}), 500
        
        logger.info(f"🔍 [API] Buscando dividendos para {ticker}...")
        dividendos = coletar_dividendos_brapi(ticker, limit=100)
        
        return jsonify({
            'status': 'ok',
            'ticker': ticker,
            'total': len(dividendos),
            'dividendos': dividendos
        })
    except Exception as e:
        logger.error(f"❌ Erro ao buscar dividendos para {ticker}: {str(e)}")
        return jsonify({'erro': str(e)}), 500

@app.route('/api/tickers')
def get_tickers():
    """Lista de tickers"""
    try:
        tickers_lista = [
            'BBSE3.SA',  # Banco do Brasil
            'CMIG4.SA',  # Cemig
            'CSMG3.SA',  # Copasa
            'ITUB4.SA',  # Itaú
            'PETR4.SA',  # Petrobras
            'SANB11.SA', # Santander
            'SYN3.SA'    # SYN
        ]
        
        if MODULOS_CARREGADOS and 'TICKERS' in globals():
            return jsonify({'tickers': TICKERS})
        else:
            return jsonify({'tickers': tickers_lista})
    except Exception as e:
        return jsonify({'tickers': tickers_lista, 'erro': str(e)})

@app.errorhandler(404)
def not_found(e):
    return jsonify({'erro': 'Endpoint não encontrado'}), 404

@app.errorhandler(500)
def internal_error(e):
    return jsonify({'erro': 'Erro interno do servidor'}), 500

if __name__ == '__main__':
    print("="*70)
    print("🚀 Dashboard Algoritimo Trade - DADOS REAIS DO MERCADO")
    print("="*70)
    print("📊 Acesse: http://localhost:5000")
    print("="*70)
    
    if MODULOS_CARREGADOS:
        print("✅ Todos os módulos carregados - Dados REAIS ativos")
    else:
        print("⚠️  Alguns módulos não carregados - Verifique dependências")
    
    print("\nPressione Ctrl+C para parar\n")
    
    try:
        app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)
    except Exception as e:
        print(f"❌ Erro: {e}")
        import traceback
        traceback.print_exc()


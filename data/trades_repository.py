import os
import sqlite3
from contextlib import contextmanager
from typing import List, Dict, Any, Tuple
from datetime import datetime
import csv
from io import StringIO, BytesIO

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "trades.db")

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    email TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    nome TEXT,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
    last_login TEXT,
    is_active INTEGER DEFAULT 1,
    is_admin INTEGER DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_users_email ON users (email);

CREATE TABLE IF NOT EXISTS trades (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL DEFAULT 1,
    trade_date TEXT NOT NULL,
    ticker TEXT NOT NULL,
    side TEXT NOT NULL CHECK (side IN ('BUY','SELL')),
    quantity REAL NOT NULL,
    price REAL NOT NULL,
    fees REAL NOT NULL DEFAULT 0,
    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_trades_ticker ON trades (ticker);
CREATE INDEX IF NOT EXISTS idx_trades_date ON trades (trade_date);
CREATE INDEX IF NOT EXISTS idx_trades_user_ticker ON trades (user_id, ticker);
CREATE INDEX IF NOT EXISTS idx_trades_user_date ON trades (user_id, trade_date);

CREATE TABLE IF NOT EXISTS dividendos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL DEFAULT 1,
    data_pagamento TEXT NOT NULL,
    data_ex_dividendo TEXT,  -- Data ex-dividendo (data a partir da qual ações são negociadas sem direito ao dividendo)
    ticker TEXT NOT NULL,
    valor_por_acao REAL NOT NULL,
    quantidade_acoes REAL NOT NULL,
    valor_total REAL NOT NULL,
    tipo TEXT DEFAULT 'DIVIDENDO' CHECK (tipo IN ('DIVIDENDO','JCP','RENDIMENTO')),
    data_busca TEXT,  -- Quando foi buscado da API
    fonte TEXT DEFAULT 'brapi.dev',  -- Fonte dos dados
    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
    UNIQUE(user_id, ticker, data_pagamento, valor_por_acao)  -- Evita duplicatas por usuário
);
CREATE INDEX IF NOT EXISTS idx_dividendos_ticker ON dividendos (ticker);
CREATE INDEX IF NOT EXISTS idx_dividendos_data ON dividendos (data_pagamento);
CREATE INDEX IF NOT EXISTS idx_dividendos_data_busca ON dividendos (data_busca);
CREATE INDEX IF NOT EXISTS idx_dividendos_data_ex ON dividendos (data_ex_dividendo);
CREATE INDEX IF NOT EXISTS idx_dividendos_user_ticker ON dividendos (user_id, ticker);
CREATE INDEX IF NOT EXISTS idx_dividendos_user_data ON dividendos (user_id, data_pagamento);
"""

@contextmanager
def _connect():
    conn = sqlite3.connect(DB_PATH)
    try:
        yield conn
    finally:
        conn.commit()
        conn.close()

def init_db() -> None:
    with _connect() as conn:
        conn.executescript(SCHEMA_SQL)

def _parse_row(row: Dict[str, str]) -> Tuple[str, str, float, float, float]:
    # Expected headers (case-insensitive): date,ticker,side,quantity,price,fees
    def g(key: str) -> str:
        for k in row.keys():
            if k.lower() == key:
                return row[k]
        return ""
    date_raw = g("date") or g("trade_date")
    ticker = (g("ticker") or "").strip().upper()
    side = (g("side") or "").strip().upper()
    quantity = float(g("quantity") or "0")
    price = float(g("price") or "0")
    fees = float(g("fees") or "0")
    # Normalize date
    # Try multiple formats
    dt = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            dt = datetime.strptime(date_raw.strip(), fmt)
            break
        except:
            continue
    if dt is None:
        # Fallback: now
        dt = datetime.now()
    trade_date = dt.strftime("%Y-%m-%d")
    # Normalize side
    if side not in ("BUY", "SELL"):
        # Accept PT terms
        if side in ("C", "COMPRA"):
            side = "BUY"
        elif side in ("V", "VENDA"):
            side = "SELL"
        else:
            side = "BUY"
    # Ensure B3 suffix handled later in pricing
    return trade_date, ticker, quantity, price, fees if side in ("BUY", "SELL") else 0.0

def import_csv(file_path: str) -> Dict[str, Any]:
    init_db()
    inserted = 0
    with open(file_path, "r", encoding="utf-8-sig") as f, _connect() as conn:
        reader = csv.DictReader(f)
        for row in reader:
            trade_date, ticker, quantity, price, fees = _parse_row(row)
            side = (row.get("side") or row.get("Side") or "BUY").strip().upper()
            if side in ("C", "COMPRA"):
                side = "BUY"
            if side in ("V", "VENDA"):
                side = "SELL"
            conn.execute(
                "INSERT INTO trades (user_id, trade_date, ticker, side, quantity, price, fees) VALUES (?,?,?,?,?,?,?)",
                (1, trade_date, ticker, side, quantity, price, fees)  # user_id=1 para compatibilidade
            )
            inserted += 1
    return {"status": "ok", "inserted": inserted}

def import_csv_bytes(csv_bytes: bytes) -> Dict[str, Any]:
    """
    Importa CSV a partir de bytes. Tenta utf-8-sig, depois latin-1.
    """
    init_db()
    inserted = 0
    text = None
    try:
        text = csv_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = csv_bytes.decode("latin-1")
    sio = StringIO(text)
    reader = csv.DictReader(sio)
    with _connect() as conn:
        for row in reader:
            trade_date, ticker, quantity, price, fees = _parse_row(row)
            side = (row.get("side") or row.get("Side") or "BUY").strip().upper()
            if side in ("C", "COMPRA"):
                side = "BUY"
            if side in ("V", "VENDA"):
                side = "SELL"
            conn.execute(
                "INSERT INTO trades (user_id, trade_date, ticker, side, quantity, price, fees) VALUES (?,?,?,?,?,?,?)",
                (1, trade_date, ticker, side, quantity, price, fees)  # user_id=1 para compatibilidade
            )
            inserted += 1
    return {"status": "ok", "inserted": inserted}

def import_excel_bytes(xlsx_bytes: bytes) -> Dict[str, Any]:
    """
    Importa XLSX (Excel) a partir de bytes usando openpyxl.
    Espera cabeçalho na primeira linha.
    """
    from openpyxl import load_workbook
    init_db()
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    # Ler cabeçalhos
    headers = []
    for cell in ws[1]:
        headers.append((cell.value or "").strip().lower())
    # Construir dicionários por linha
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for idx, val in enumerate(r):
            key = headers[idx] if idx < len(headers) else f"col{idx}"
            row_dict[key] = val if val is not None else ""
        rows.append(row_dict)
    return insert_rows(rows)

def insert_rows(rows: List[Dict[str, Any]], user_id: int = None) -> Dict[str, Any]:
    """Insere trades. user_id é obrigatório para multi-tenant."""
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    inserted = 0
    with _connect() as conn:
        for row in rows:
            # Obter data e converter Timestamp do pandas para string se necessário
            trade_date_raw = row.get("trade_date") or row.get("date") or datetime.now()
            
            # Converter Timestamp do pandas ou datetime para string
            if hasattr(trade_date_raw, 'strftime'):
                # É um datetime ou Timestamp do pandas
                trade_date = trade_date_raw.strftime("%Y-%m-%d")
            elif isinstance(trade_date_raw, str):
                # Já é string, tentar parsear e normalizar
                try:
                    from datetime import datetime as dt
                    # Tentar diferentes formatos
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y"]:
                        try:
                            dt_obj = dt.strptime(trade_date_raw.strip(), fmt)
                            trade_date = dt_obj.strftime("%Y-%m-%d")
                            break
                        except:
                            continue
                    else:
                        # Se não conseguiu parsear, usar como está
                        trade_date = trade_date_raw.strip()
                except:
                    trade_date = trade_date_raw.strip()
            else:
                # Fallback: usar data atual
                trade_date = datetime.now().strftime("%Y-%m-%d")
            
            ticker = (row.get("ticker") or "").strip().upper()
            side = (row.get("side") or "BUY").strip().upper()
            quantity = float(row.get("quantity") or 0)
            price = float(row.get("price") or 0)
            fees = float(row.get("fees") or 0)
            conn.execute(
                "INSERT INTO trades (user_id, trade_date, ticker, side, quantity, price, fees) VALUES (?,?,?,?,?,?,?)",
                (user_id, trade_date, ticker, side, quantity, price, fees)
            )
            inserted += 1
    return {"status": "ok", "inserted": inserted}

def list_trades(limit: int = 200, user_id: int = None) -> List[Dict[str, Any]]:
    """Lista trades. user_id é obrigatório para multi-tenant."""
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        cur = conn.execute(
            "SELECT id, trade_date, ticker, side, quantity, price, fees FROM trades WHERE user_id = ? ORDER BY trade_date DESC, id DESC LIMIT ?", 
            (user_id, limit)
        )
        rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            "id": r[0],
            "trade_date": r[1],
            "ticker": r[2],
            "side": r[3],
            "quantity": r[4],
            "price": r[5],
            "fees": r[6],
        })
    return result

def positions_summary(user_id: int = None) -> Dict[str, Any]:
    """
    Consolida posições por ticker usando método de preço médio móvel:
    - BUY: novo_preco_medio = (qtd*pm + qtd_buy*preco) / (qtd + qtd_buy)
    - SELL: reduz a quantidade; preço médio permanece para a posição remanescente
    Ignora PnL realizado aqui; foco é posição aberta (qty e preço médio).
    Também rastreia a data da primeira compra ainda aberta para cálculo de rentabilidade anualizada.
    user_id é obrigatório para multi-tenant.
    """
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        cur = conn.execute(
            "SELECT ticker, side, quantity, price, fees, trade_date, id FROM trades WHERE user_id = ? ORDER BY ticker ASC, trade_date ASC, id ASC",
            (user_id,)
        )
        rows = cur.fetchall()
    per_ticker = {}
    for ticker, side, qty, price, fees, trade_date, _id in rows:
        ticker = (ticker or "").upper().strip()
        if not ticker:
            continue
        state = per_ticker.get(ticker, {"qty": 0.0, "avg": 0.0, "fees": 0.0, "first_buy_date": None})
        side = (side or "BUY").upper()
        q = float(qty or 0)
        p = float(price or 0)
        f = float(fees or 0)
        if side == "BUY":
            if q > 0:
                new_qty = state["qty"] + q
                if new_qty > 0:
                    state["avg"] = ((state["qty"] * state["avg"]) + (q * p)) / new_qty
                state["qty"] = new_qty
                # Rastrear data da primeira compra (quando posição estava zerada e começou a comprar)
                if state["first_buy_date"] is None and state["qty"] > 0:
                    state["first_buy_date"] = trade_date
        elif side == "SELL":
            if q > 0:
                state["qty"] = state["qty"] - q
                if state["qty"] < 1e-9:
                    state["qty"] = 0.0
                    state["avg"] = 0.0
                    state["first_buy_date"] = None  # Reset quando posição zera
        state["fees"] += f
        per_ticker[ticker] = state
    result = []
    for ticker, st in per_ticker.items():
        if abs(st["qty"]) < 1e-9:
            continue
        result.append({
            "ticker": ticker,
            "net_quantity": float(st["qty"]),
            "avg_cost": float(st["avg"]),
            "fees_total": float(st["fees"]),
            "first_buy_date": st["first_buy_date"]
        })
    return {"positions": sorted(result, key=lambda x: x["ticker"])}

def calculate_realized_pnl(user_id: int = None) -> Dict[str, Any]:
    """
    Calcula o PnL realizado (lucro/prejuízo das vendas já executadas).
    Usa método FIFO: as primeiras compras são as primeiras vendas.
    Retorna: total_pnl_realizado, total_custo_vendas, total_receita_vendas
    user_id é obrigatório para multi-tenant.
    """
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        cur = conn.execute(
            "SELECT ticker, side, quantity, price, fees, trade_date, id FROM trades WHERE user_id = ? ORDER BY ticker ASC, trade_date ASC, id ASC",
            (user_id,)
        )
        rows = cur.fetchall()
    
    # Para cada ticker, manter fila de compras (FIFO)
    per_ticker = {}
    total_pnl_realizado = 0.0
    total_custo_vendas = 0.0
    total_receita_vendas = 0.0
    
    for ticker, side, qty, price, fees, trade_date, _id in rows:
        ticker = (ticker or "").upper().strip()
        if not ticker:
            continue
        
        side = (side or "BUY").upper()
        q = float(qty or 0)
        p = float(price or 0)
        f = float(fees or 0)
        
        if ticker not in per_ticker:
            per_ticker[ticker] = {"buys": []}  # Fila de compras: [(qty, price, date, id), ...]
        
        if side == "BUY":
            if q > 0:
                per_ticker[ticker]["buys"].append((q, p, trade_date, _id))
        elif side == "SELL":
            if q > 0:
                # Processar venda usando FIFO
                remaining_sell = q
                while remaining_sell > 1e-9 and len(per_ticker[ticker]["buys"]) > 0:
                    buy_qty, buy_price, buy_date, buy_id = per_ticker[ticker]["buys"][0]
                    
                    # Calcular quanto dessa compra foi vendido
                    qty_sold = min(remaining_sell, buy_qty)
                    
                    # Calcular PnL dessa venda parcial
                    custo = qty_sold * buy_price
                    receita = qty_sold * p
                    pnl = receita - custo
                    
                    total_pnl_realizado += pnl
                    total_custo_vendas += custo
                    total_receita_vendas += receita
                    
                    # Atualizar fila de compras
                    remaining_sell -= qty_sold
                    if buy_qty - qty_sold < 1e-9:
                        # Compra totalmente consumida
                        per_ticker[ticker]["buys"].pop(0)
                    else:
                        # Compra parcialmente consumida
                        per_ticker[ticker]["buys"][0] = (buy_qty - qty_sold, buy_price, buy_date, buy_id)
    
    return {
        "total_pnl_realizado": total_pnl_realizado,
        "total_custo_vendas": total_custo_vendas,
        "total_receita_vendas": total_receita_vendas
    }

def reset_trades(user_id: int = None) -> Dict[str, Any]:
    """
    Remove todas as operações da base (truncate lógico).
    user_id é obrigatório para multi-tenant.
    """
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        conn.execute("DELETE FROM trades WHERE user_id = ?", (user_id,))
    return {"status": "ok", "deleted": True}

# ========== FUNÇÕES DE DIVIDENDOS ==========

def _parse_dividendo_row(row: Dict[str, str]) -> Tuple[str, str, float, float, float, str]:
    """Parseia uma linha de dividendo. Espera: data_pagamento, ticker, valor_por_acao, quantidade_acoes, tipo (opcional)"""
    def g(key: str) -> str:
        for k in row.keys():
            if k.lower() == key.lower():
                return row[k]
        return ""
    
    date_raw = g("data_pagamento") or g("data") or g("date")
    ticker = (g("ticker") or "").strip().upper()
    valor_por_acao = float(g("valor_por_acao") or g("valor") or "0")
    quantidade_acoes = float(g("quantidade_acoes") or g("quantidade") or g("qty") or "0")
    tipo = (g("tipo") or "DIVIDENDO").strip().upper()
    
    if tipo not in ("DIVIDENDO", "JCP", "RENDIMENTO"):
        tipo = "DIVIDENDO"
    
    # Normalizar data
    dt = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            dt = datetime.strptime(date_raw.strip(), fmt)
            break
        except:
            continue
    if dt is None:
        dt = datetime.now()
    data_pagamento = dt.strftime("%Y-%m-%d")
    
    valor_total = valor_por_acao * quantidade_acoes
    
    return data_pagamento, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo

def import_dividendos_csv_bytes(csv_bytes: bytes) -> Dict[str, Any]:
    """Importa dividendos a partir de bytes CSV."""
    init_db()
    inserted = 0
    text = None
    try:
        text = csv_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = csv_bytes.decode("latin-1")
    sio = StringIO(text)
    reader = csv.DictReader(sio)
    with _connect() as conn:
        for row in reader:
            data_pagamento, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo = _parse_dividendo_row(row)
            conn.execute(
                "INSERT INTO dividendos (data_pagamento, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo) VALUES (?,?,?,?,?,?)",
                (data_pagamento, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo)
            )
            inserted += 1
    return {"status": "ok", "inserted": inserted}

def import_dividendos_excel_bytes(xlsx_bytes: bytes) -> Dict[str, Any]:
    """Importa dividendos a partir de bytes Excel."""
    from openpyxl import load_workbook
    init_db()
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        headers.append((cell.value or "").strip().lower())
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for idx, val in enumerate(r):
            key = headers[idx] if idx < len(headers) else f"col{idx}"
            row_dict[key] = val if val is not None else ""
        rows.append(row_dict)
    return insert_dividendos_rows(rows)

def insert_dividendos_rows(rows: List[Dict[str, Any]], fonte: str = "brapi.dev", user_id: int = None) -> Dict[str, Any]:
    """
    Insere dividendos a partir de lista de dicionários.
    Usa INSERT OR IGNORE para evitar duplicatas (devido ao UNIQUE constraint).
    user_id é obrigatório para multi-tenant.
    """
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    inserted = 0
    skipped = 0
    data_busca = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def _normalize_date(date_value):
        """Normaliza data de Timestamp do pandas ou datetime para string"""
        if date_value is None:
            return datetime.now().strftime("%Y-%m-%d")
        
        if hasattr(date_value, 'strftime'):
            # É um datetime ou Timestamp do pandas
            return date_value.strftime("%Y-%m-%d")
        elif isinstance(date_value, str):
            # Já é string, tentar parsear e normalizar
            try:
                from datetime import datetime as dt
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y"]:
                    try:
                        dt_obj = dt.strptime(date_value.strip(), fmt)
                        return dt_obj.strftime("%Y-%m-%d")
                    except:
                        continue
                return date_value.strip()
            except:
                return date_value.strip()
        else:
            return datetime.now().strftime("%Y-%m-%d")
    
    with _connect() as conn:
        for row in rows:
            data_pagamento_raw = row.get("data_pagamento") or row.get("data") or row.get("date") or datetime.now()
            data_pagamento = _normalize_date(data_pagamento_raw)
            
            data_ex_dividendo_raw = row.get("data_ex_dividendo") or data_pagamento_raw
            data_ex_dividendo = _normalize_date(data_ex_dividendo_raw)
            
            ticker = (row.get("ticker") or "").strip().upper()
            valor_por_acao = float(row.get("valor_por_acao") or row.get("valor") or 0)
            quantidade_acoes = float(row.get("quantidade_acoes") or row.get("quantidade") or row.get("qty") or 0)
            tipo = (row.get("tipo") or "DIVIDENDO").strip().upper()
            if tipo not in ("DIVIDENDO", "JCP", "RENDIMENTO"):
                tipo = "DIVIDENDO"
            valor_total = valor_por_acao * quantidade_acoes
            
            # Usar INSERT OR IGNORE para evitar duplicatas
            cursor = conn.execute(
                """INSERT OR IGNORE INTO dividendos 
                   (user_id, data_pagamento, data_ex_dividendo, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo, data_busca, fonte) 
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (user_id, data_pagamento, data_ex_dividendo, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo, data_busca, fonte)
            )
            if cursor.rowcount > 0:
                inserted += 1
            else:
                skipped += 1
    return {"status": "ok", "inserted": inserted, "skipped": skipped}

def calcular_quantidade_acoes_na_data(ticker: str, data: str, user_id: int = None) -> float:
    """
    Calcula a quantidade de ações que o usuário tinha em uma data específica.
    Usa FIFO para determinar a posição histórica.
    
    Args:
        ticker: Código da ação
        data: Data no formato 'YYYY-MM-DD'
        user_id: ID do usuário (obrigatório para multi-tenant)
    
    Returns:
        Quantidade de ações naquela data (pode ser 0 se não tinha ações)
    """
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        # Buscar todas as operações até a data especificada
        cur = conn.execute(
            """SELECT side, quantity, trade_date, id 
               FROM trades 
               WHERE user_id = ? AND ticker = ? AND trade_date <= ?
               ORDER BY trade_date ASC, id ASC""",
            (user_id, ticker, data)
        )
        rows = cur.fetchall()
    
    quantidade = 0.0
    for side, qty, trade_date, _id in rows:
        side = (side or "BUY").upper()
        q = float(qty or 0)
        
        if side == "BUY":
            quantidade += q
        elif side == "SELL":
            quantidade -= q
            if quantidade < 0:
                quantidade = 0.0  # Não pode ter quantidade negativa
    
    return max(0.0, quantidade)

def verificar_necessidade_sincronizacao_dividendos(ticker: str, horas_cache: int = 24, user_id: int = None) -> bool:
    """
    Verifica se precisa sincronizar dividendos de um ticker.
    Retorna True se não há dados ou se os dados são mais antigos que horas_cache.
    user_id é obrigatório para multi-tenant.
    """
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        # Buscar último dividendo e data da última busca
        cur = conn.execute(
            """SELECT MAX(data_busca) as ultima_busca, COUNT(*) as total 
               FROM dividendos WHERE user_id = ? AND ticker = ?""",
            (user_id, ticker)
        )
        row = cur.fetchone()
        ultima_busca = row[0] if row and row[0] else None
        total = row[1] if row else 0
        
        if not ultima_busca or total == 0:
            return True  # Precisa buscar
        
        # Verificar se ultima busca é mais antiga que horas_cache
        try:
            from datetime import timedelta
            ultima_busca_dt = datetime.strptime(ultima_busca, "%Y-%m-%d %H:%M:%S")
            agora = datetime.now()
            diferenca = agora - ultima_busca_dt
            return diferenca.total_seconds() > (horas_cache * 3600)
        except:
            return True  # Em caso de erro, busca novamente

def list_dividendos(limit: int = 200, user_id: int = None) -> List[Dict[str, Any]]:
    """Lista dividendos recentes. user_id é obrigatório para multi-tenant."""
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        cur = conn.execute(
            "SELECT id, data_pagamento, data_ex_dividendo, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo FROM dividendos WHERE user_id = ? ORDER BY data_pagamento DESC, id DESC LIMIT ?",
            (user_id, limit)
        )
        rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            "id": r[0],
            "data_pagamento": r[1],
            "data_ex_dividendo": r[2],
            "ticker": r[3],
            "valor_por_acao": r[4],
            "quantidade_acoes": r[5],
            "valor_total": r[6],
            "tipo": r[7]
        })
    return result

def list_dividendos_por_ticker(ticker: str, user_id: int = None) -> List[Dict[str, Any]]:
    """Lista dividendos recebidos para um ticker específico. user_id é obrigatório para multi-tenant."""
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        cur = conn.execute(
            """SELECT id, data_pagamento, data_ex_dividendo, ticker, valor_por_acao, quantidade_acoes, valor_total, tipo 
               FROM dividendos 
               WHERE user_id = ? AND ticker = ? 
               ORDER BY data_pagamento DESC, id DESC""",
            (user_id, ticker)
        )
        rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            "id": r[0],
            "data_pagamento": r[1],
            "data_ex_dividendo": r[2],
            "ticker": r[3],
            "valor_por_acao": r[4],
            "quantidade_acoes": r[5],
            "valor_total": r[6],
            "tipo": r[7]
        })
    return result

def calculate_total_dividendos(user_id: int = None) -> Dict[str, Any]:
    """Calcula o total de dividendos recebidos (por ticker e total geral). user_id é obrigatório para multi-tenant."""
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        # Total geral
        cur = conn.execute("SELECT SUM(valor_total) FROM dividendos WHERE user_id = ?", (user_id,))
        total_geral = cur.fetchone()[0] or 0.0
        
        # Por ticker
        cur = conn.execute(
            "SELECT ticker, SUM(valor_total) as total FROM dividendos WHERE user_id = ? GROUP BY ticker ORDER BY ticker",
            (user_id,)
        )
        rows = cur.fetchall()
    
    por_ticker = {}
    for ticker, total in rows:
        por_ticker[ticker] = float(total or 0)
    
    return {
        "total_geral": float(total_geral),
        "por_ticker": por_ticker
    }

def reset_dividendos(user_id: int = None) -> Dict[str, Any]:
    """Remove todos os dividendos da base. user_id é obrigatório para multi-tenant."""
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    init_db()
    with _connect() as conn:
        conn.execute("DELETE FROM dividendos WHERE user_id = ?", (user_id,))
    return {"status": "ok", "deleted": True}

def limpar_dividendos_invalidos(user_id: int = None) -> Dict[str, Any]:
    """
    Remove dividendos que têm data ex-dividendo anterior à primeira compra do papel.
    Este agente verifica todos os dividendos no banco e remove aqueles que não deveriam
    estar lá porque a compra foi feita depois da data ex-dividendo.
    
    Returns:
        Dicionário com estatísticas da limpeza
    """
    import logging
    logger = logging.getLogger(__name__)
    
    init_db()
    
    logger.info("🧹 [LIMPEZA] Iniciando limpeza de dividendos inválidos...")
    
    total_removidos = 0
    total_verificados = 0
    removidos_por_ticker = {}
    
    if user_id is None:
        user_id = 1  # Fallback para compatibilidade
    
    with _connect() as conn:
        # Buscar todos os tickers únicos que têm dividendos para este usuário
        cur = conn.execute("SELECT DISTINCT ticker FROM dividendos WHERE user_id = ?", (user_id,))
        tickers_com_dividendos = [row[0] for row in cur.fetchall()]
        
        logger.info(f"🔍 [LIMPEZA] Encontrados {len(tickers_com_dividendos)} tickers com dividendos para user_id={user_id}: {tickers_com_dividendos}")
        
        for ticker in tickers_com_dividendos:
            try:
                # Buscar primeira data de compra do ticker para este usuário
                cur = conn.execute(
                    "SELECT MIN(trade_date) FROM trades WHERE user_id = ? AND ticker = ? AND side = 'BUY'",
                    (user_id, ticker)
                )
                row = cur.fetchone()
                primeira_compra = row[0] if row and row[0] else None
                
                if not primeira_compra:
                    logger.warning(f"⚠️  [LIMPEZA] {ticker}: Nenhuma compra encontrada. Removendo todos os dividendos deste ticker.")
                    # Se não há compras, remover todos os dividendos deste ticker para este usuário
                    cur = conn.execute("SELECT COUNT(*) FROM dividendos WHERE user_id = ? AND ticker = ?", (user_id, ticker))
                    count = cur.fetchone()[0] or 0
                    if count > 0:
                        conn.execute("DELETE FROM dividendos WHERE user_id = ? AND ticker = ?", (user_id, ticker))
                        removidos_por_ticker[ticker] = count
                        total_removidos += count
                        logger.info(f"  ✅ {ticker}: {count} dividendos removidos (sem compras)")
                    continue
                
                logger.info(f"📅 [LIMPEZA] {ticker}: Primeira compra em {primeira_compra}")
                
                # Buscar dividendos deste ticker para este usuário
                cur = conn.execute(
                    """SELECT id, data_pagamento, data_ex_dividendo 
                       FROM dividendos 
                       WHERE user_id = ? AND ticker = ?""",
                    (user_id, ticker)
                )
                dividendos = cur.fetchall()
                
                removidos_ticker = 0
                from datetime import datetime as dt
                
                # Converter primeira compra para datetime uma vez
                try:
                    primeira_compra_dt = dt.strptime(primeira_compra, "%Y-%m-%d").date()
                except Exception as e:
                    logger.error(f"  ❌ {ticker}: Erro ao parsear primeira compra {primeira_compra}: {str(e)[:100]}")
                    continue
                
                for div_id, data_pagamento, data_ex_dividendo in dividendos:
                    total_verificados += 1
                    
                    # Usar data_ex_dividendo se disponível, senão usar data_pagamento como fallback
                    data_referencia = data_ex_dividendo if data_ex_dividendo else data_pagamento
                    
                    if not data_referencia:
                        # Se não tem nenhuma data, remover
                        conn.execute("DELETE FROM dividendos WHERE user_id = ? AND id = ?", (user_id, div_id))
                        removidos_ticker += 1
                        logger.debug(f"  🗑️  {ticker}: Dividendo {div_id} removido (sem data)")
                        continue
                    
                    # Comparar datas usando datetime para garantir comparação correta
                    try:
                        # Tentar parsear a data de referência
                        data_ref_dt = None
                        for fmt in ["%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"]:
                            try:
                                data_ref_dt = dt.strptime(data_referencia, fmt).date()
                                break
                            except:
                                continue
                        
                        if data_ref_dt is None:
                            logger.warning(f"  ⚠️  {ticker}: Não foi possível parsear data {data_referencia} para dividendo {div_id}")
                            continue
                        
                        # Comparar: se data_ex_dividendo < primeira_compra, remover
                        # REGRA: Para receber dividendo, precisa ter comprado ANTES da data ex-dividendo
                        # Se comprou na ou depois da data ex-dividendo, não tem direito
                        if data_ref_dt < primeira_compra_dt:
                            conn.execute("DELETE FROM dividendos WHERE user_id = ? AND id = ?", (user_id, div_id))
                            removidos_ticker += 1
                            logger.info(f"  🗑️  {ticker}: Dividendo {data_pagamento} (ex: {data_ex_dividendo or 'N/A'}) removido - data ex ({data_ref_dt}) é anterior à primeira compra ({primeira_compra_dt})")
                        elif data_ref_dt == primeira_compra_dt:
                            # Se comprou no mesmo dia da data ex-dividendo, também não tem direito (comprou depois do fechamento)
                            conn.execute("DELETE FROM dividendos WHERE user_id = ? AND id = ?", (user_id, div_id))
                            removidos_ticker += 1
                            logger.info(f"  🗑️  {ticker}: Dividendo {data_pagamento} (ex: {data_ex_dividendo or 'N/A'}) removido - compra ({primeira_compra_dt}) foi no mesmo dia da data ex-dividendo")
                    except Exception as e_parse:
                        logger.warning(f"  ⚠️  {ticker}: Erro ao comparar datas para dividendo {div_id}: {str(e_parse)[:100]}")
                        # Em caso de erro, não remover (mais seguro)
                        continue
                
                if removidos_ticker > 0:
                    removidos_por_ticker[ticker] = removidos_ticker
                    total_removidos += removidos_ticker
                    logger.info(f"  ✅ {ticker}: {removidos_ticker} dividendos inválidos removidos")
                else:
                    logger.info(f"  ✅ {ticker}: Nenhum dividendo inválido encontrado")
                    
            except Exception as e:
                logger.error(f"  ❌ {ticker}: Erro ao processar: {str(e)[:100]}")
                continue
    
    logger.info(f"✅ [LIMPEZA] Limpeza concluída:")
    logger.info(f"   - Tickers verificados: {len(tickers_com_dividendos)}")
    logger.info(f"   - Dividendos verificados: {total_verificados}")
    logger.info(f"   - Dividendos removidos: {total_removidos}")
    
    return {
        "status": "ok",
        "tickers_verificados": len(tickers_com_dividendos),
        "total_verificados": total_verificados,
        "total_removidos": total_removidos,
        "removidos_por_ticker": removidos_por_ticker
    }


